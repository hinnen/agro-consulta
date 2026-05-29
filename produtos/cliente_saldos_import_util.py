"""Leitura e casamento de planilhas ERP → saldos em ClienteAgro."""

from __future__ import annotations

import csv
import re
import unicodedata
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Literal

from produtos.models import ClienteAgro

TipoSaldoImport = Literal["auto", "cashback", "vale"]
ModoImport = Literal["substituir", "somar"]

_COL_NOME = (
    "nome",
    "cliente",
    "nome cliente",
    "nome do cliente",
    "razao social",
    "razaosocial",
    "nome fantasia",
    "nomefantasia",
    "pessoa",
    "favorecido",
)
_COL_VALOR = ("valor", "vlr", "montante", "total")
_COL_CASHBACK = (
    "saldo_cashback",
    "saldo cashback",
    "cashback",
    "valor cashback",
    "credito cashback",
    "crédito cashback",
    "saldo atual",
    "saldo disponivel",
    "saldo disponível",
)
_COL_VALE = (
    "saldo_vale_credito",
    "saldo vale",
    "saldo vale credito",
    "saldo vale crédito",
    "vale credito",
    "vale crédito",
    "vale",
    "credito vale",
    "crédito vale",
)
_COL_FIADO = (
    "limite_fiado_local",
    "limite fiado",
    "limite de fiado",
    "limite credito",
    "limite crédito",
)


def _norm_header(s: str) -> str:
    t = unicodedata.normalize("NFKD", str(s or ""))
    t = "".join(ch for ch in t if not unicodedata.combining(ch))
    t = t.lower().strip()
    t = re.sub(r"\s+", " ", t)
    return t


def _norm_nome(s: str) -> str:
    t = _norm_header(s)
    return t.upper()


def _norm_nome_match_key(s: str) -> str:
    """
    Chave para casar planilha ERP ↔ ClienteAgro (ignora prefixos tipo «031 », «Sn - »).
    """
    t = _norm_nome(s)
    t = re.sub(r"^\d{1,5}\s+", "", t)
    t = re.sub(r"^[A-Z]{1,6}\s*-\s*", "", t)
    return t.strip()


def _dec(v) -> Decimal:
    txt = str(v or "").strip()
    if not txt or txt in ("-", "—"):
        return Decimal("0")
    txt = txt.replace("R$", "").replace(" ", "")
    if "," in txt:
        txt = txt.replace(".", "").replace(",", ".")
    try:
        return Decimal(txt).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _pick_col(fieldnames: list[str], aliases: tuple[str, ...]) -> str | None:
    norm_map = {_norm_header(h): h for h in fieldnames if h}
    for alias in aliases:
        if alias in norm_map:
            return norm_map[alias]
    for h in fieldnames:
        nh = _norm_header(h)
        for alias in aliases:
            if alias in nh or nh in alias:
                return h
    return None


def _inferir_tipo_saldo_pelo_arquivo(path: Path | None) -> TipoSaldoImport | None:
    if path is None:
        return None
    stem = _norm_header(path.stem)
    if "vale" in stem and "cashback" not in stem:
        return "vale"
    if "cashback" in stem:
        return "cashback"
    return None


def detectar_colunas(
    fieldnames: list[str],
    *,
    tipo_saldo: TipoSaldoImport = "auto",
    path: Path | None = None,
) -> dict[str, str | None]:
    if tipo_saldo == "auto":
        inferido = _inferir_tipo_saldo_pelo_arquivo(path)
        if inferido:
            tipo_saldo = inferido
    headers_norm = " ".join(_norm_header(h) for h in fieldnames if h)
    if tipo_saldo == "auto":
        if "cashback" in headers_norm and "vale" not in headers_norm:
            tipo_saldo = "cashback"
        elif "vale" in headers_norm and "cashback" not in headers_norm:
            tipo_saldo = "vale"
    nome = _pick_col(fieldnames, _COL_NOME)
    cashback = _pick_col(fieldnames, _COL_CASHBACK)
    vale = _pick_col(fieldnames, _COL_VALE)
    fiado = _pick_col(fieldnames, _COL_FIADO)
    valor_gen = _pick_col(fieldnames, _COL_VALOR)

    if cashback and valor_gen and cashback == valor_gen and not vale:
        if tipo_saldo == "vale":
            cashback = None
        elif tipo_saldo in ("cashback", "auto"):
            vale = None

    if tipo_saldo == "vale" and not vale:
        vale = valor_gen or cashback
        cashback = None
    elif tipo_saldo == "cashback" and not cashback:
        cashback = valor_gen or vale
        vale = None
    elif tipo_saldo == "auto":
        if not cashback and not vale and valor_gen:
            cashback = valor_gen
        elif cashback and vale and cashback == vale:
            cashback = valor_gen
            vale = None

    # Planilha Codigo + Cliente + Valor (3 colunas)
    if nome and not cashback and not vale and len(fieldnames) == 3:
        for h in fieldnames:
            if h != nome and _norm_header(h) not in ("codigo", "cod", "id", "código"):
                if tipo_saldo == "vale":
                    vale = h
                else:
                    cashback = h
                break

    if nome and not cashback and not vale and len(fieldnames) == 2:
        outra = fieldnames[0] if fieldnames[1] == nome else fieldnames[1]
        if outra != nome:
            if tipo_saldo == "vale":
                vale = outra
            else:
                cashback = outra

    return {
        "nome": nome,
        "saldo_cashback": cashback,
        "saldo_vale_credito": vale,
        "limite_fiado_local": fiado,
    }


def _ler_xlsx(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return [], []
        headers = [str(c or "").strip() for c in header_row]
        rows: list[dict[str, Any]] = []
        for raw in rows_iter:
            if raw is None:
                continue
            if not any(v is not None and str(v).strip() for v in raw):
                continue
            row = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                row[h] = raw[i] if i < len(raw) else ""
            rows.append(row)
        return headers, rows
    finally:
        wb.close()


def _ler_csv(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            with path.open("r", encoding=enc, newline="") as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    return [], []
                headers = list(reader.fieldnames)
                return headers, list(reader)
        except UnicodeDecodeError:
            continue
    raise ValueError(f"Não foi possível ler o CSV com encodings comuns: {path}")


def ler_planilha(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    suf = path.suffix.lower()
    if suf == ".csv":
        return _ler_csv(path)
    if suf in (".xlsx", ".xlsm"):
        return _ler_xlsx(path)
    raise ValueError("Use arquivo .csv ou .xlsx")


def buscar_clientes_por_nome(nome: str) -> list[ClienteAgro]:
    chave = _norm_nome_match_key(nome)
    if len(chave) < 2:
        return []
    candidatos: list[ClienteAgro] = []
    for c in ClienteAgro.objects.filter(ativo=True).only(
        "pk", "nome", "saldo_cashback", "saldo_vale_credito", "limite_fiado_local", "editado_local"
    ):
        if _norm_nome_match_key(c.nome) == chave:
            candidatos.append(c)
    if candidatos:
        return candidatos
    # Fallback: mesma chave contida no nome (um único candidato).
    parciais: list[ClienteAgro] = []
    for c in ClienteAgro.objects.filter(ativo=True).only(
        "pk", "nome", "saldo_cashback", "saldo_vale_credito", "limite_fiado_local", "editado_local"
    ):
        cn = _norm_nome_match_key(c.nome)
        if chave in cn or cn in chave:
            parciais.append(c)
    if len(parciais) == 1:
        return parciais
    return []


def _agregar_por_cliente(
    rows: list[dict[str, Any]],
    cols: dict[str, str | None],
) -> tuple[list[dict[str, Any]], int]:
    """
    Soma saldos quando o mesmo cliente aparece em várias linhas (ex.: vale crédito ERP).
    Retorna linhas agregadas e quantidade de linhas extras fundidas.
    """
    buckets: dict[str, dict[str, Any]] = {}
    extras = 0

    for row in rows:
        nome_raw = str(row.get(cols["nome"]) or "").strip()
        if not nome_raw:
            continue
        key = _norm_nome(nome_raw)
        if not key:
            continue

        cb = _dec(row.get(cols["saldo_cashback"])) if cols.get("saldo_cashback") else Decimal("0")
        vl = _dec(row.get(cols["saldo_vale_credito"])) if cols.get("saldo_vale_credito") else Decimal("0")
        fd = _dec(row.get(cols["limite_fiado_local"])) if cols.get("limite_fiado_local") else Decimal("0")

        if key not in buckets:
            buckets[key] = {
                "nome_planilha": nome_raw,
                "saldo_cashback": cb if cols.get("saldo_cashback") else None,
                "saldo_vale_credito": vl if cols.get("saldo_vale_credito") else None,
                "limite_fiado_local": fd if cols.get("limite_fiado_local") else None,
                "linhas": 1,
            }
            continue

        extras += 1
        b = buckets[key]
        b["linhas"] = int(b.get("linhas") or 1) + 1
        if cols.get("saldo_cashback"):
            b["saldo_cashback"] = (b.get("saldo_cashback") or Decimal("0")) + cb
        if cols.get("saldo_vale_credito"):
            b["saldo_vale_credito"] = (b.get("saldo_vale_credito") or Decimal("0")) + vl
        if cols.get("limite_fiado_local"):
            b["limite_fiado_local"] = (b.get("limite_fiado_local") or Decimal("0")) + fd

    return list(buckets.values()), extras


def aplicar_importacao(
    path: Path,
    *,
    dry_run: bool = False,
    tipo_saldo: TipoSaldoImport = "auto",
    modo: ModoImport = "substituir",
    col_nome: str | None = None,
    col_cashback: str | None = None,
    col_vale: str | None = None,
    col_fiado: str | None = None,
    relatorio_path: Path | None = None,
) -> dict[str, Any]:
    headers, rows = ler_planilha(path)
    if not headers:
        raise ValueError("Planilha sem cabeçalho.")

    cols = detectar_colunas(headers, tipo_saldo=tipo_saldo, path=path)
    if col_nome:
        cols["nome"] = col_nome
    if col_cashback:
        cols["saldo_cashback"] = col_cashback
    if col_vale:
        cols["saldo_vale_credito"] = col_vale
    if col_fiado:
        cols["limite_fiado_local"] = col_fiado

    if not cols.get("nome"):
        raise ValueError(
            f"Coluna de nome não encontrada. Cabeçalhos: {headers}. "
            "Use --col-nome se o ERP usar outro título."
        )
    if not any(cols.get(k) for k in ("saldo_cashback", "saldo_vale_credito", "limite_fiado_local")):
        raise ValueError(
            f"Nenhuma coluna de saldo detectada. Cabeçalhos: {headers}. "
            "Use --tipo vale --col-vale Valor (planilha Codigo/Cliente/Valor)."
        )

    agregados, linhas_fundidas = _agregar_por_cliente(rows, cols)

    alterados = 0
    ignorados = sum(
        1 for row in rows if not str(row.get(cols["nome"]) or "").strip()
    )
    nao_encontrados = 0
    ambiguos = 0
    relatorio: list[dict[str, str]] = []

    for item in agregados:
        nome_raw = str(item.get("nome_planilha") or "").strip()
        matches = buscar_clientes_por_nome(nome_raw)
        if not matches:
            nao_encontrados += 1
            relatorio.append({"nome_planilha": nome_raw, "status": "nao_encontrado"})
            continue
        if len(matches) > 1:
            ambiguos += 1
            relatorio.append(
                {
                    "nome_planilha": nome_raw,
                    "status": f"ambiguo_{len(matches)}",
                }
            )
            continue

        cli = matches[0]
        upd: dict[str, Decimal] = {}
        if cols.get("saldo_cashback") and item.get("saldo_cashback") is not None:
            novo = _dec(item.get("saldo_cashback"))
            upd["saldo_cashback"] = (
                _dec(cli.saldo_cashback) + novo if modo == "somar" else novo
            )
        if cols.get("saldo_vale_credito") and item.get("saldo_vale_credito") is not None:
            novo = _dec(item.get("saldo_vale_credito"))
            upd["saldo_vale_credito"] = (
                _dec(cli.saldo_vale_credito) + novo if modo == "somar" else novo
            )
        if cols.get("limite_fiado_local") and item.get("limite_fiado_local") is not None:
            novo = _dec(item.get("limite_fiado_local"))
            upd["limite_fiado_local"] = (
                _dec(cli.limite_fiado_local) + novo if modo == "somar" else novo
            )

        if not dry_run:
            from django.utils import timezone as tz

            # Só saldos — não marca editado_local (preserva sync ERP/Mongo no cadastro).
            ClienteAgro.objects.filter(pk=cli.pk).update(
                **{k: v.quantize(Decimal("0.01")) for k, v in upd.items()},
                atualizado_em=tz.now(),
            )
        alterados += 1

    if relatorio_path and relatorio:
        relatorio_path.parent.mkdir(parents=True, exist_ok=True)
        with relatorio_path.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["nome_planilha", "status"])
            w.writeheader()
            w.writerows(relatorio)

    if alterados and not dry_run:
        try:
            from django.core.cache import cache

            from produtos.views import API_LIST_CUSTOMERS_CACHE_KEY

            cache.delete(API_LIST_CUSTOMERS_CACHE_KEY)
        except Exception:
            pass

    return {
        "alterados": alterados,
        "ignorados": ignorados,
        "nao_encontrados": nao_encontrados,
        "ambiguos": ambiguos,
        "linhas_lidas": len(rows),
        "clientes_unicos": len(agregados),
        "linhas_fundidas": linhas_fundidas,
        "colunas_usadas": cols,
        "tipo_saldo": tipo_saldo,
        "modo": modo,
        "dry_run": dry_run,
        "relatorio": str(relatorio_path) if relatorio_path else "",
    }

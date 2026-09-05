"""Exportação / importação Excel do cadastro de produtos (fase 1 — overlay Agro)."""

from __future__ import annotations

import re
import unicodedata
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any

from django.core.cache import cache
from django.db import transaction
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from produtos.models import CadastroPlanilhaImportHistoricoAgro, ProdutoGestaoOverlayAgro

EXPORT_MAX_ROWS = 15000
IMPORT_MAX_ROWS = 2500

COL_ID = "id"
COL_CODIGO_GM = "codigo_gm"
COL_NOME = "nome"
COL_MARCA = "marca"
COL_MODELO = "modelo"
COL_UNIDADE = "unidade"
COL_PESO = "peso_etiqueta"
COL_CATEGORIA = "categoria"
COL_SUBCATEGORIA = "subcategoria"
COL_SUBCATEGORIA_2 = "subcategoria_2"
COL_SUBCATEGORIA_3 = "subcategoria_3"
COL_SUBCATEGORIA_4 = "subcategoria_4"
COL_CODIGO_BARRAS = "codigo_barras"
COL_PRECO_CUSTO = "preco_custo"
COL_PRECO_VENDA = "preco_venda"

# Aba Delivery / catálogo público (`cadastro_extras.delivery`)
COL_DEL_ATIVO = "delivery_ativo"
COL_DEL_TITULO = "delivery_titulo"
COL_DEL_DESCRICAO = "delivery_descricao"
COL_DEL_ORDEM = "delivery_ordem"
COL_DEL_DESTAQUE = "delivery_destaque"
COL_DEL_ESTOQUE_NEG = "delivery_estoque_negativo"
COL_DEL_PESO = "delivery_peso"
COL_DEL_CAT = "delivery_categoria"
COL_DEL_SUB1 = "delivery_sub1"
COL_DEL_SUB2 = "delivery_sub2"
COL_DEL_SUB3 = "delivery_sub3"
COL_DEL_SUB4 = "delivery_sub4"
COL_DEL_EMBALAGENS = "delivery_embalagens"

# Só Excel ↓ (Entrada NF) — import ignora
COL_FORN_COMPRA_1 = "fornecedor_compra_1"
COL_FORN_COMPRA_2 = "fornecedor_compra_2"
COL_FORN_COMPRA_3 = "fornecedor_compra_3"

FORNECEDOR_EXPORT_KEYS = (
    COL_FORN_COMPRA_1,
    COL_FORN_COMPRA_2,
    COL_FORN_COMPRA_3,
)

DELIVERY_IMPORT_KEYS = (
    COL_DEL_ATIVO,
    COL_DEL_TITULO,
    COL_DEL_DESCRICAO,
    COL_DEL_ORDEM,
    COL_DEL_DESTAQUE,
    COL_DEL_ESTOQUE_NEG,
    COL_DEL_PESO,
    COL_DEL_CAT,
    COL_DEL_SUB1,
    COL_DEL_SUB2,
    COL_DEL_SUB3,
    COL_DEL_SUB4,
    COL_DEL_EMBALAGENS,
)

EXPORT_HEADERS: list[tuple[str, str]] = [
    ("ID", COL_ID),
    ("Código GM", COL_CODIGO_GM),
    ("Nome", COL_NOME),
    ("Marca", COL_MARCA),
    ("Categoria", COL_CATEGORIA),
    ("Subcategoria", COL_SUBCATEGORIA),
    ("Subcategoria 2", COL_SUBCATEGORIA_2),
    ("Subcategoria 3", COL_SUBCATEGORIA_3),
    ("Subcategoria 4", COL_SUBCATEGORIA_4),
    ("Unidade", COL_UNIDADE),
    ("Modelo", COL_MODELO),
    ("Peso", COL_PESO),
    ("Código barras", COL_CODIGO_BARRAS),
    ("Preço custo", COL_PRECO_CUSTO),
    ("Preço venda", COL_PRECO_VENDA),
    ("Delivery ativo", COL_DEL_ATIVO),
    ("Delivery título", COL_DEL_TITULO),
    ("Delivery descrição", COL_DEL_DESCRICAO),
    ("Delivery ordem", COL_DEL_ORDEM),
    ("Delivery destaque", COL_DEL_DESTAQUE),
    ("Delivery estoque neg.", COL_DEL_ESTOQUE_NEG),
    ("Delivery peso", COL_DEL_PESO),
    ("Delivery categoria", COL_DEL_CAT),
    ("Delivery sub 1", COL_DEL_SUB1),
    ("Delivery sub 2", COL_DEL_SUB2),
    ("Delivery sub 3", COL_DEL_SUB3),
    ("Delivery sub 4", COL_DEL_SUB4),
    ("Delivery embalagens", COL_DEL_EMBALAGENS),
    ("Últ. fornecedor", COL_FORN_COMPRA_1),
    ("2º fornecedor", COL_FORN_COMPRA_2),
    ("3º fornecedor", COL_FORN_COMPRA_3),
]

_COLS_TEXTO_EXCEL = frozenset(
    {
        COL_ID,
        COL_CODIGO_GM,
        COL_CODIGO_BARRAS,
        COL_UNIDADE,
        COL_MODELO,
        COL_PESO,
        COL_SUBCATEGORIA,
        COL_SUBCATEGORIA_2,
        COL_SUBCATEGORIA_3,
        COL_SUBCATEGORIA_4,
        COL_DEL_ATIVO,
        COL_DEL_TITULO,
        COL_DEL_DESCRICAO,
        COL_DEL_ORDEM,
        COL_DEL_DESTAQUE,
        COL_DEL_ESTOQUE_NEG,
        COL_DEL_PESO,
        COL_DEL_CAT,
        COL_DEL_SUB1,
        COL_DEL_SUB2,
        COL_DEL_SUB3,
        COL_DEL_SUB4,
        COL_DEL_EMBALAGENS,
        COL_FORN_COMPRA_1,
        COL_FORN_COMPRA_2,
        COL_FORN_COMPRA_3,
    }
)

_MAX_TXT_IMPORT = {
    COL_NOME: 300,
    COL_CODIGO_GM: 64,
    COL_CODIGO_BARRAS: 80,
    COL_CATEGORIA: 200,
    COL_SUBCATEGORIA: 200,
    COL_SUBCATEGORIA_2: 200,
    COL_SUBCATEGORIA_3: 200,
    COL_SUBCATEGORIA_4: 200,
    COL_UNIDADE: 20,
    COL_PESO: 40,
    COL_MODELO: 200,
    COL_MARCA: 120,
    COL_DEL_TITULO: 200,
    COL_DEL_DESCRICAO: 2000,
    COL_DEL_PESO: 40,
    COL_DEL_CAT: 80,
    COL_DEL_SUB1: 80,
    COL_DEL_SUB2: 80,
    COL_DEL_SUB3: 80,
    COL_DEL_SUB4: 80,
    COL_DEL_EMBALAGENS: 500,
    COL_DEL_ORDEM: 10,
    COL_DEL_ATIVO: 10,
    COL_DEL_DESTAQUE: 10,
    COL_DEL_ESTOQUE_NEG: 10,
}

# Colunas travadas no Excel (só ID — coluna oculta na planilha).
EXPORT_COLS_BLOQUEADAS = frozenset({COL_ID})
EXPORT_COLS_OCULTAS = frozenset({COL_ID})

EXPORT_COL_KEYS = [key for _, key in EXPORT_HEADERS]

def normalizar_colunas_export(raw: str | None) -> list[str]:
    """Colunas pedidas na exportação — ID sempre incluído."""
    if not raw or not str(raw).strip():
        return list(EXPORT_COL_KEYS)
    pedidas = [x.strip() for x in str(raw).split(",") if x.strip()]
    out: list[str] = []
    for k in pedidas:
        if k in EXPORT_COL_KEYS and k not in out:
            out.append(k)
    if COL_ID not in out:
        out.insert(0, COL_ID)
    return out or list(EXPORT_COL_KEYS)


def normalizar_categorias_export(raw: str | None) -> list[str]:
    if not raw or not str(raw).strip():
        return []
    return list(dict.fromkeys(x.strip() for x in str(raw).split("|") if x.strip()))[:80]


def _filtrar_rows_categorias(rows: list[dict], categorias: list[str]) -> list[dict]:
    if not categorias:
        return rows
    alvo = {c.strip().lower() for c in categorias if c.strip()}
    if not alvo:
        return rows
    return [r for r in rows if str(r.get("categoria") or "").strip().lower() in alvo]


_COLS_FACETA = (
    COL_MARCA,
    COL_CATEGORIA,
    COL_SUBCATEGORIA,
    COL_SUBCATEGORIA_2,
    COL_SUBCATEGORIA_3,
    COL_SUBCATEGORIA_4,
    COL_UNIDADE,
)
_FACETA_ROTULO = {
    COL_MARCA: "Marca",
    COL_CATEGORIA: "Categoria",
    COL_SUBCATEGORIA: "Subcategoria",
    COL_SUBCATEGORIA_2: "Subcategoria 2",
    COL_SUBCATEGORIA_3: "Subcategoria 3",
    COL_SUBCATEGORIA_4: "Subcategoria 4",
    COL_UNIDADE: "Unidade",
}
_FACETA_FUZZY_MIN = 0.86
_FACETA_FUZZY_MAX_LEN_DIFF = 3


def _norm_faceta_chave(s: str) -> str:
    t = unicodedata.normalize("NFD", str(s or ""))
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", t).strip().lower()


def carregar_facetas_planilha() -> dict[str, list[str]]:
    """Listas conhecidas para validação de marca/categoria/sub na importação."""
    from produtos.agro_fonte_config import agro_catalogo_usa_postgres

    marcas: list[str] = []
    categorias: list[str] = []
    subcategorias: list[str] = []
    unidades: list[str] = []

    try:
        if agro_catalogo_usa_postgres():
            from produtos import catalogo_agro

            fac = catalogo_agro.facetas_gestao(limite=2000)
            marcas = list(fac.get("marcas") or [])
            categorias = list(fac.get("categorias") or [])
            subcategorias = list(fac.get("subcategorias") or [])
            unidades = list(fac.get("unidades") or [])
            qs = catalogo_agro.queryset_catalogo_ativos(inativos=True)
            for fld in ("subcategoria_2", "subcategoria_3", "subcategoria_4"):
                subcategorias.extend(
                    str(x).strip()
                    for x in qs.exclude(**{fld: ""}).values_list(fld, flat=True).distinct()[:800]
                    if str(x or "").strip()
                )
    except Exception:
        pass

    ov_qs = ProdutoGestaoOverlayAgro.objects.all()
    marcas.extend(
        str(x).strip()
        for x in ov_qs.exclude(marca="").values_list("marca", flat=True).distinct()[:2000]
        if str(x or "").strip()
    )
    categorias.extend(
        str(x).strip()
        for x in ov_qs.exclude(categoria="").values_list("categoria", flat=True).distinct()[:1200]
        if str(x or "").strip()
    )
    unidades.extend(
        str(x).strip()
        for x in ov_qs.exclude(unidade="").values_list("unidade", flat=True).distinct()[:400]
        if str(x or "").strip()
    )
    for fld in ("subcategoria", "subcategoria_2", "subcategoria_3", "subcategoria_4"):
        subcategorias.extend(
            str(x).strip()
            for x in ov_qs.exclude(**{fld: ""}).values_list(fld, flat=True).distinct()[:800]
            if str(x or "").strip()
        )

    def _uniq(vals: list[str], lim: int = 1500) -> list[str]:
        seen: set[str] = set()
        out: list[str] = []
        for v in vals:
            s = str(v or "").strip()
            if not s:
                continue
            k = _norm_faceta_chave(s)
            if k in seen:
                continue
            seen.add(k)
            out.append(s)
            if len(out) >= lim:
                break
        out.sort(key=lambda x: x.lower())
        return out

    return {
        COL_MARCA: _uniq(marcas, 2500),
        COL_CATEGORIA: _uniq(categorias, 1200),
        COL_SUBCATEGORIA: _uniq(subcategorias, 1500),
        COL_SUBCATEGORIA_2: _uniq(subcategorias, 1500),
        COL_SUBCATEGORIA_3: _uniq(subcategorias, 1500),
        COL_SUBCATEGORIA_4: _uniq(subcategorias, 1500),
        COL_UNIDADE: _uniq(unidades, 400),
    }


def _mapa_canonico_faceta(lista: list[str]) -> dict[str, str]:
    return {_norm_faceta_chave(v): v for v in lista if str(v or "").strip()}


def _sugerir_faceta(valor: str, conhecidos: list[str]) -> tuple[str | None, float]:
    from difflib import SequenceMatcher

    alvo = _norm_faceta_chave(valor)
    if not alvo or not conhecidos:
        return None, 0.0
    melhor = None
    melhor_score = 0.0
    alvo_len = len(alvo)
    for cand in conhecidos:
        ck = _norm_faceta_chave(cand)
        if not ck:
            continue
        if abs(len(ck) - alvo_len) > _FACETA_FUZZY_MAX_LEN_DIFF:
            continue
        sc = SequenceMatcher(None, alvo, ck).ratio()
        if sc > melhor_score:
            melhor_score = sc
            melhor = cand
            if sc >= 0.99:
                break
    if melhor and melhor_score >= _FACETA_FUZZY_MIN:
        return melhor, float(melhor_score)
    return None, float(melhor_score)


def _classificar_valor_faceta(
    campo: str,
    valor: str,
    facetas: dict[str, list[str]],
    canonicos: dict[str, dict[str, str]],
) -> dict[str, Any]:
    raw = str(valor or "").strip()
    rotulo = _FACETA_ROTULO.get(campo, campo)
    if not raw:
        return {"campo": campo, "rotulo": rotulo, "valor": "", "status": "vazio"}
    canon_map = canonicos.get(campo) or {}
    chave = _norm_faceta_chave(raw)
    if chave in canon_map:
        return {
            "campo": campo,
            "rotulo": rotulo,
            "valor": raw,
            "valor_final": canon_map[chave],
            "status": "ok",
            "score": 1.0,
        }
    sug, score = _sugerir_faceta(raw, facetas.get(campo) or [])
    if sug:
        return {
            "campo": campo,
            "rotulo": rotulo,
            "valor": raw,
            "valor_final": sug,
            "sugestao": sug,
            "status": "corrigir",
            "score": round(score, 3),
        }
    return {
        "campo": campo,
        "rotulo": rotulo,
        "valor": raw,
        "valor_final": raw,
        "status": "novo",
        "score": round(score, 3) if score else 0.0,
    }


def _resolver_facetas_no_patch(
    patch: dict[str, Any],
    facetas: dict[str, list[str]],
    canonicos: dict[str, dict[str, str]],
    *,
    permitir_novos: bool,
    linha: int | None = None,
) -> tuple[dict[str, Any], list[dict], list[str]]:
    out = dict(patch)
    eventos: list[dict] = []
    erros: list[str] = []
    for campo in _COLS_FACETA:
        if campo not in out:
            continue
        info = _classificar_valor_faceta(campo, str(out[campo] or ""), facetas, canonicos)
        info["linha"] = linha
        st = info.get("status")
        if st == "ok":
            out[campo] = info["valor_final"]
            if info["valor_final"] != str(patch[campo] or "").strip():
                eventos.append({**info, "acao": "canonico"})
        elif st == "corrigir":
            out[campo] = info["valor_final"]
            eventos.append({**info, "acao": "corrigir"})
        elif st == "novo":
            eventos.append({**info, "acao": "novo"})
            if not permitir_novos:
                erros.append(
                    f"{info['rotulo']} «{info['valor']}» não cadastrado"
                    + (f" (parecido: {info.get('sugestao')})" if info.get("sugestao") else "")
                    + ". Marque «Permitir criar novos» ou use um nome da lista."
                )
    return out, eventos, erros


def _resumir_eventos_faceta(eventos: list[dict]) -> tuple[list[dict], list[dict]]:
    novos_map: dict[tuple[str, str], dict] = {}
    corr_map: dict[tuple[str, str], dict] = {}
    for ev in eventos:
        campo = str(ev.get("campo") or "")
        valor = str(ev.get("valor") or "")
        key = (campo, _norm_faceta_chave(valor))
        linha = ev.get("linha")
        if ev.get("status") == "novo" or ev.get("acao") == "novo":
            slot = novos_map.setdefault(
                key,
                {
                    "campo": campo,
                    "rotulo": ev.get("rotulo") or campo,
                    "valor": valor,
                    "linhas": [],
                    "score": ev.get("score"),
                },
            )
            if linha and linha not in slot["linhas"]:
                slot["linhas"].append(linha)
        elif ev.get("status") == "corrigir" or ev.get("acao") == "corrigir":
            slot = corr_map.setdefault(
                key,
                {
                    "campo": campo,
                    "rotulo": ev.get("rotulo") or campo,
                    "valor": valor,
                    "sugestao": ev.get("sugestao") or ev.get("valor_final"),
                    "score": ev.get("score"),
                    "linhas": [],
                },
            )
            if linha and linha not in slot["linhas"]:
                slot["linhas"].append(linha)
    return list(novos_map.values()), list(corr_map.values())


def headers_export(colunas: list[str] | None = None) -> list[tuple[str, str]]:
    cols = colunas or list(EXPORT_COL_KEYS)
    return [(label, key) for label, key in EXPORT_HEADERS if key in cols]


IMPORT_KEYS = {
    COL_CODIGO_GM,
    COL_NOME,
    COL_MARCA,
    COL_MODELO,
    COL_UNIDADE,
    COL_PESO,
    COL_CATEGORIA,
    COL_SUBCATEGORIA,
    COL_SUBCATEGORIA_2,
    COL_SUBCATEGORIA_3,
    COL_SUBCATEGORIA_4,
    COL_CODIGO_BARRAS,
    COL_PRECO_CUSTO,
    COL_PRECO_VENDA,
    *DELIVERY_IMPORT_KEYS,
}

OVERLAY_IMPORT_KEYS = (
    COL_CODIGO_GM,
    COL_NOME,
    COL_MARCA,
    COL_MODELO,
    COL_UNIDADE,
    COL_PESO,
    COL_CATEGORIA,
    COL_SUBCATEGORIA,
    COL_SUBCATEGORIA_2,
    COL_SUBCATEGORIA_3,
    COL_SUBCATEGORIA_4,
    COL_CODIGO_BARRAS,
    COL_PRECO_VENDA,
    *DELIVERY_IMPORT_KEYS,
)

HISTORICO_IMPORT_LISTA_LIMITE = 30


def _overlay_model_field(key: str) -> str:
    return "codigo_nfe" if key == COL_CODIGO_GM else key


def _max_txt_import(key: str) -> int:
    return _MAX_TXT_IMPORT.get(key, 120)


def _extras_dict(ov: ProdutoGestaoOverlayAgro | None) -> dict:
    if ov is None:
        return {}
    return dict(ov.cadastro_extras) if isinstance(getattr(ov, "cadastro_extras", None), dict) else {}


def _ler_overlay_import_campo(ov: ProdutoGestaoOverlayAgro | None, key: str):
    if ov is None:
        return None if key == COL_PRECO_VENDA else ""
    if key == COL_PRECO_VENDA:
        val = ov.preco_venda
        return str(val) if val is not None else None
    if key == COL_MODELO:
        return str(_extras_dict(ov).get("modelo") or "")
    if key in DELIVERY_IMPORT_KEYS:
        return _delivery_planilha_de_ov(ov).get(key, "")
    return str(getattr(ov, _overlay_model_field(key), "") or "")


def _gravar_overlay_import_campo(ov: ProdutoGestaoOverlayAgro, key: str, val) -> None:
    if key == COL_PRECO_VENDA:
        if val is None or (isinstance(val, str) and not str(val).strip()):
            ov.preco_venda = None
        else:
            ov.preco_venda = Decimal(str(val))
        return
    if key == COL_MODELO:
        ex = _extras_dict(ov)
        s = str(val or "").strip()[:200]
        if s:
            ex["modelo"] = s
        else:
            ex.pop("modelo", None)
        ov.cadastro_extras = ex
        return
    if key in DELIVERY_IMPORT_KEYS:
        # Gravação em lote via _aplicar_patch_delivery (evita reescrever N vezes).
        return
    setattr(ov, _overlay_model_field(key), str(val or "")[: _max_txt_import(key)])


def _sim_nao(flag: bool) -> str:
    return "Sim" if flag else "Não"


def _parse_bool_planilha(val) -> bool | None:
    s = _cel_str(val).strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    if s in ("1", "s", "sim", "true", "yes", "y", "x", "ativo", "on"):
        return True
    if s in ("0", "n", "nao", "false", "no", "inativo", "off"):
        return False
    return None


def _eh_limpar_planilha(val: str) -> bool:
    s = (val or "").strip().lower()
    return s in ("-", ".", "limpar", "apagar", "none", "null")


def _mapa_cats_delivery() -> dict[int, Any]:
    from produtos.models import CatalogoDeliveryCategoria

    return {c.pk: c for c in CatalogoDeliveryCategoria.objects.all()}


def _nome_cat_delivery(mapa: dict[int, Any], cid: int) -> str:
    if not cid:
        return ""
    c = mapa.get(int(cid))
    return (c.nome if c else "") or ""


def _delivery_planilha_de_dict(d: dict, *, mapa_cats: dict[int, Any] | None = None) -> dict[str, str]:
    from produtos.catalogo_delivery_util import normalizar_delivery

    d = normalizar_delivery(d or {})
    ids_cat = [
        int(d.get("categoria_id") or 0),
        int(d.get("subcategoria_id") or 0),
        int(d.get("subcategoria2_id") or 0),
        int(d.get("subcategoria3_id") or 0),
        int(d.get("subcategoria4_id") or 0),
    ]
    if mapa_cats is not None:
        mapa = mapa_cats
    elif any(ids_cat):
        mapa = _mapa_cats_delivery()
    else:
        mapa = {}
    emb_parts: list[str] = []
    for e in d.get("embalagens") or []:
        pid = str(e.get("produto_id") or "").strip()
        if not pid:
            continue
        rot = str(e.get("rotulo") or "").strip()
        cod = _codigo_gm_de_pid(pid) or pid
        emb_parts.append(f"{cod}:{rot}" if rot else cod)
    return {
        COL_DEL_ATIVO: _sim_nao(bool(d.get("ativo"))),
        COL_DEL_TITULO: str(d.get("titulo") or ""),
        COL_DEL_DESCRICAO: str(d.get("descricao") or ""),
        COL_DEL_ORDEM: str(int(d.get("ordem") or 0)),
        COL_DEL_DESTAQUE: _sim_nao(bool(d.get("destaque"))),
        COL_DEL_ESTOQUE_NEG: _sim_nao(bool(d.get("permitir_estoque_negativo"))),
        COL_DEL_PESO: str(d.get("peso_texto") or ""),
        COL_DEL_CAT: _nome_cat_delivery(mapa, ids_cat[0]),
        COL_DEL_SUB1: _nome_cat_delivery(mapa, ids_cat[1]),
        COL_DEL_SUB2: _nome_cat_delivery(mapa, ids_cat[2]),
        COL_DEL_SUB3: _nome_cat_delivery(mapa, ids_cat[3]),
        COL_DEL_SUB4: _nome_cat_delivery(mapa, ids_cat[4]),
        COL_DEL_EMBALAGENS: " | ".join(emb_parts),
    }


def _delivery_planilha_de_ov(ov: ProdutoGestaoOverlayAgro | None) -> dict[str, str]:
    from produtos.catalogo_delivery_util import delivery_de_extras

    if ov is None:
        return _delivery_planilha_de_dict({})
    return _delivery_planilha_de_dict(delivery_de_extras(_extras_dict(ov)))


def _codigo_gm_de_pid(pid: str) -> str:
    pid = str(pid or "").strip()[:64]
    if not pid:
        return ""
    ov = ProdutoGestaoOverlayAgro.objects.filter(produto_externo_id=pid).only("codigo_nfe").first()
    if ov and (ov.codigo_nfe or "").strip():
        return (ov.codigo_nfe or "").strip()[:64]
    try:
        from produtos.models import Produto

        p = Produto.objects.filter(produto_externo_id=pid).only("codigo_nfe").first()
        if p and (p.codigo_nfe or "").strip():
            return (p.codigo_nfe or "").strip()[:64]
    except Exception:
        pass
    return ""


def _pid_de_codigo_gm_ou_id(token: str) -> str:
    t = str(token or "").strip()[:64]
    if not t:
        return ""
    if _id_produto_planilha_valido(t) and (
        re.fullmatch(r"[0-9a-f]{24}", t.lower())
        or t.upper().startswith("AGRO")
        or t.isdigit()
    ):
        if ProdutoGestaoOverlayAgro.objects.filter(produto_externo_id=t).exists():
            return t
        try:
            from produtos.models import Produto

            if Produto.objects.filter(produto_externo_id=t).exists():
                return t
        except Exception:
            pass
    ov = (
        ProdutoGestaoOverlayAgro.objects.filter(codigo_nfe__iexact=t)
        .only("produto_externo_id")
        .first()
    )
    if ov and ov.produto_externo_id:
        return str(ov.produto_externo_id)[:64]
    try:
        from produtos.models import Produto

        p = Produto.objects.filter(codigo_nfe__iexact=t).only("produto_externo_id").first()
        if p and p.produto_externo_id:
            return str(p.produto_externo_id)[:64]
    except Exception:
        pass
    return ""


def _resolver_filho_categoria(parent_id: int | None, nome: str) -> tuple[int, str | None]:
    from produtos.models import CatalogoDeliveryCategoria

    nome = (nome or "").strip()
    if not nome:
        return 0, None
    qs = CatalogoDeliveryCategoria.objects.filter(ativo=True, nome__iexact=nome)
    if parent_id:
        qs = qs.filter(parent_id=parent_id)
    else:
        qs = qs.filter(parent__isnull=True)
    hit = qs.order_by("ordem", "pk").first()
    if not hit:
        nivel = "categoria" if not parent_id else "subcategoria"
        return 0, f"«{nome}» não encontrada como {nivel} no catálogo Delivery."
    return int(hit.pk), None


def _resolver_caminho_delivery_nomes(
    nomes: list[str],
) -> tuple[dict[str, int], str | None]:
    """nomes = [cat, sub1, sub2, sub3, sub4] — vazio = zera daí em diante se veio na lista."""
    ids = [0, 0, 0, 0, 0]
    parent: int | None = None
    for i, nome in enumerate(nomes[:5]):
        n = (nome or "").strip()
        if not n or _eh_limpar_planilha(n):
            for j in range(i, 5):
                ids[j] = 0
            break
        cid, err = _resolver_filho_categoria(parent, n)
        if err:
            return {}, err
        ids[i] = cid
        parent = cid
    return {
        "categoria_id": ids[0],
        "subcategoria_id": ids[1],
        "subcategoria2_id": ids[2],
        "subcategoria3_id": ids[3],
        "subcategoria4_id": ids[4],
    }, None


def _parse_embalagens_planilha(raw: str) -> tuple[list[dict], str | None]:
    s = str(raw or "").strip()
    if not s or _eh_limpar_planilha(s):
        return [], None
    parts = re.split(r"[|;,]+", s)
    out: list[dict] = []
    seen: set[str] = set()
    for part in parts:
        part = part.strip()
        if not part:
            continue
        if ":" in part:
            cod, rot = part.split(":", 1)
            cod, rot = cod.strip(), rot.strip()[:40]
        else:
            cod, rot = part, ""
        pid = _pid_de_codigo_gm_ou_id(cod)
        if not pid:
            return [], f"Embalagem «{cod}» não encontrada (use Código GM ou ID)."
        if pid in seen:
            continue
        seen.add(pid)
        out.append({"produto_id": pid, "rotulo": rot})
        if len(out) >= 6:
            break
    return out, None


def _enriquecer_row_delivery_planilha(
    row: dict,
    ov: ProdutoGestaoOverlayAgro | None = None,
    *,
    mapa_cats: dict[int, Any] | None = None,
) -> None:
    from produtos.catalogo_delivery_util import delivery_de_extras

    if ov is None:
        fields = _delivery_planilha_de_dict({}, mapa_cats=mapa_cats)
    else:
        fields = _delivery_planilha_de_dict(
            delivery_de_extras(_extras_dict(ov)), mapa_cats=mapa_cats
        )
    row.update(fields)


def _enriquecer_rows_delivery_batch(rows: list[dict], ovs: dict) -> None:
    if not rows:
        return
    mapa = _mapa_cats_delivery()
    for r in rows:
        pid = str(r.get("id") or "")
        _enriquecer_row_delivery_planilha(r, ovs.get(pid), mapa_cats=mapa)


def _aplicar_patch_delivery(ov: ProdutoGestaoOverlayAgro, patch: dict) -> str | None:
    """Aplica colunas Delivery no overlay. Retorna mensagem de erro ou None."""
    from produtos.catalogo_delivery_util import delivery_de_extras, normalizar_delivery

    if not any(k in patch for k in DELIVERY_IMPORT_KEYS):
        return None
    d = dict(delivery_de_extras(_extras_dict(ov)))
    atual_pl = _delivery_planilha_de_dict(d)

    if COL_DEL_ATIVO in patch:
        b = _parse_bool_planilha(patch[COL_DEL_ATIVO])
        if b is None:
            return "Delivery ativo: use Sim ou Não."
        d["ativo"] = b
    if COL_DEL_TITULO in patch:
        v = str(patch[COL_DEL_TITULO] or "").strip()
        d["titulo"] = "" if _eh_limpar_planilha(v) else v[:200]
    if COL_DEL_DESCRICAO in patch:
        v = str(patch[COL_DEL_DESCRICAO] or "").strip()
        d["descricao"] = "" if _eh_limpar_planilha(v) else v[:2000]
    if COL_DEL_ORDEM in patch:
        try:
            d["ordem"] = max(0, min(9999, int(str(patch[COL_DEL_ORDEM]).strip() or "0")))
        except (TypeError, ValueError):
            return "Delivery ordem: número inválido."
    if COL_DEL_DESTAQUE in patch:
        b = _parse_bool_planilha(patch[COL_DEL_DESTAQUE])
        if b is None:
            return "Delivery destaque: use Sim ou Não."
        d["destaque"] = b
    if COL_DEL_ESTOQUE_NEG in patch:
        b = _parse_bool_planilha(patch[COL_DEL_ESTOQUE_NEG])
        if b is None:
            return "Delivery estoque neg.: use Sim ou Não."
        d["permitir_estoque_negativo"] = b
    if COL_DEL_PESO in patch:
        v = str(patch[COL_DEL_PESO] or "").strip()
        d["peso_texto"] = "" if _eh_limpar_planilha(v) else v[:40]

    path_cols = (COL_DEL_CAT, COL_DEL_SUB1, COL_DEL_SUB2, COL_DEL_SUB3, COL_DEL_SUB4)
    if any(k in patch for k in path_cols):
        nomes = [
            str(patch[k]) if k in patch else atual_pl.get(k, "")
            for k in path_cols
        ]
        # Se coluna veio no patch como limpar, força limpeza a partir daí
        for i, k in enumerate(path_cols):
            if k in patch and _eh_limpar_planilha(str(patch[k] or "")):
                for j in range(i, 5):
                    nomes[j] = "-"
                break
        ids, err = _resolver_caminho_delivery_nomes(nomes)
        if err:
            return err
        d.update(ids)

    if COL_DEL_EMBALAGENS in patch:
        emb, err = _parse_embalagens_planilha(str(patch[COL_DEL_EMBALAGENS] or ""))
        if err:
            return err
        d["embalagens"] = emb

    d_norm = normalizar_delivery(d, processar_imagem=False)
    ex = _extras_dict(ov)
    if d_norm.get("ativo") or any(
        (
            d_norm.get("titulo"),
            d_norm.get("descricao"),
            d_norm.get("imagem_base64"),
            d_norm.get("peso_texto"),
            d_norm.get("permitir_estoque_negativo"),
            d_norm.get("destaque"),
            int(d_norm.get("ordem") or 0) > 0,
            int(d_norm.get("categoria_id") or 0) > 0,
            int(d_norm.get("subcategoria_id") or 0) > 0,
            int(d_norm.get("subcategoria2_id") or 0) > 0,
            int(d_norm.get("subcategoria3_id") or 0) > 0,
            int(d_norm.get("subcategoria4_id") or 0) > 0,
            bool(d_norm.get("embalagens")),
        )
    ):
        ex["delivery"] = d_norm
    else:
        ex.pop("delivery", None)
    ov.cadastro_extras = ex
    return None


def _validar_patch_delivery(atual: dict, patch: dict) -> str | None:
    """Valida Delivery sem gravar (overlay temporário)."""
    if not any(k in patch for k in DELIVERY_IMPORT_KEYS):
        return None
    pid = str(atual.get("id") or "").strip()[:64]
    ov = ProdutoGestaoOverlayAgro(produto_externo_id=pid or "tmp")
    real = ProdutoGestaoOverlayAgro.objects.filter(produto_externo_id=pid).first() if pid else None
    if real and isinstance(real.cadastro_extras, dict):
        ov.cadastro_extras = dict(real.cadastro_extras)
    else:
        ov.cadastro_extras = {}
    return _aplicar_patch_delivery(ov, patch)


# Produto PG: só os campos que a planilha mexeu (evita zerar Unidade ao gravar só Modelo/Peso/Sub 2).
_PG_TXT_FROM_PATCH = (
    (COL_NOME, "nome", 300),
    (COL_MARCA, "marca", 120),
    (COL_MODELO, "modelo", 200),
    (COL_UNIDADE, "unidade", 20),
    (COL_CATEGORIA, "categoria", 200),
    (COL_SUBCATEGORIA, "subcategoria", 200),
    (COL_SUBCATEGORIA_2, "subcategoria_2", 200),
    (COL_SUBCATEGORIA_3, "subcategoria_3", 200),
    (COL_SUBCATEGORIA_4, "subcategoria_4", 200),
    (COL_CODIGO_BARRAS, "codigo_barras", 50),
    (COL_CODIGO_GM, "codigo_nfe", 64),
)


def _snapshot_produto_pg(p_pg) -> dict[str, Any]:
    return {
        "custo": float(p_pg.custo or 0),
        "preco_venda": float(p_pg.preco_venda or 0),
        "nome": (p_pg.nome or "")[:300],
        "marca": (p_pg.marca or "")[:120],
        "modelo": str(getattr(p_pg, "modelo", None) or "")[:200],
        "unidade": (p_pg.unidade or "")[:20],
        "categoria": (p_pg.categoria or "")[:200],
        "subcategoria": (p_pg.subcategoria or "")[:200],
        "subcategoria_2": (p_pg.subcategoria_2 or "")[:200],
        "subcategoria_3": (p_pg.subcategoria_3 or "")[:200],
        "subcategoria_4": (p_pg.subcategoria_4 or "")[:200],
        "codigo_barras": (p_pg.codigo_barras or "")[:50],
        "codigo_nfe": (p_pg.codigo_nfe or "")[:64],
    }


def _valor_overlay_para_produto(ov, col: str):
    if col == COL_MODELO:
        return str(_extras_dict(ov).get("modelo") or "")[:200]
    if col == COL_CODIGO_GM:
        return (ov.codigo_nfe or "").strip()[:64]
    if col == COL_CODIGO_BARRAS:
        return (ov.codigo_barras or "").strip()[:50] or None
    if col == COL_UNIDADE:
        return ((ov.unidade or "").strip() or "UN")[:20]
    if col == COL_CATEGORIA:
        return (ov.categoria or "").strip()[:200] or None
    if col == COL_NOME:
        return (ov.nome or "").strip()[:300]
    if col == COL_MARCA:
        return (ov.marca or "").strip()[:120]
    if col == COL_SUBCATEGORIA:
        return (ov.subcategoria or "").strip()[:200]
    if col == COL_SUBCATEGORIA_2:
        return (ov.subcategoria_2 or "").strip()[:200]
    if col == COL_SUBCATEGORIA_3:
        return (ov.subcategoria_3 or "").strip()[:200]
    if col == COL_SUBCATEGORIA_4:
        return (ov.subcategoria_4 or "").strip()[:200]
    return None


def _aplicar_patch_no_produto_pg(pid: str, ov, patch: dict, custo_payload=None) -> None:
    from produtos import catalogo_agro

    p = catalogo_agro.obter_produto_model(pid)
    if p is None:
        sync_payload = {}
        if COL_MODELO in patch:
            sync_payload["modelo"] = str(_extras_dict(ov).get("modelo") or "")
        catalogo_agro.sincronizar_modelo_produto_de_overlay(
            pid, ov, custo_payload=custo_payload, payload=sync_payload or None
        )
        return
    changed = False
    for col, attr, _mx in _PG_TXT_FROM_PATCH:
        if col not in patch:
            continue
        setattr(p, attr, _valor_overlay_para_produto(ov, col))
        changed = True
    if COL_PRECO_VENDA in patch and ov.preco_venda is not None:
        p.preco_venda = ov.preco_venda
        changed = True
    if custo_payload is not None:
        p.custo = custo_payload
        changed = True
    if changed:
        p.save()


def _restaurar_produto_pg(p, pg_antes: dict, patch_keys) -> bool:
    changed = False
    if COL_PRECO_CUSTO in patch_keys and "custo" in pg_antes:
        p.custo = Decimal(str(pg_antes["custo"])).quantize(Decimal("0.01"))
        changed = True
    if COL_PRECO_VENDA in patch_keys and "preco_venda" in pg_antes:
        p.preco_venda = Decimal(str(pg_antes["preco_venda"])).quantize(Decimal("0.01"))
        changed = True
    for col, attr, mx in _PG_TXT_FROM_PATCH:
        if col not in patch_keys or attr not in pg_antes:
            continue
        val = str(pg_antes.get(attr) or "")[:mx]
        if attr == "unidade":
            val = val or "UN"
        elif attr == "codigo_barras":
            val = val or None
        elif attr == "categoria":
            val = val or None
        setattr(p, attr, val)
        changed = True
    return changed


def _valor_atual_campo_import(atual: dict, key: str):
    if key in (COL_PRECO_CUSTO, COL_PRECO_VENDA):
        return atual.get(key)
    if key == COL_CODIGO_GM:
        return str(atual.get("codigo_gm") or atual.get("codigo_nfe") or "").strip()
    return str(atual.get(key) or "").strip()


def _norm_header(h: str) -> str:
    s = unicodedata.normalize("NFD", str(h or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.lower().strip()
    s = re.sub(r"\s+", " ", s)
    return s


def _map_headers(headers: list[str]) -> dict[str, str | None]:
    norm = {_norm_header(h): h for h in headers if str(h or "").strip()}
    aliases: dict[str, tuple[str, ...]] = {
        COL_ID: ("id", "produto id", "produto_id", "codigo produto"),
        COL_CODIGO_GM: ("codigo gm", "codigo_gm", "codigo nfe", "codigo erp", "gm"),
        COL_NOME: ("nome", "produto", "descricao produto"),
        COL_MARCA: ("marca",),
        COL_CATEGORIA: ("categoria", "grupo"),
        COL_SUBCATEGORIA: ("subcategoria", "sub categoria", "subgrupo"),
        COL_SUBCATEGORIA_2: ("subcategoria 2", "sub categoria 2", "sub 2", "sub2"),
        COL_SUBCATEGORIA_3: ("subcategoria 3", "sub categoria 3", "sub 3", "sub3"),
        COL_SUBCATEGORIA_4: ("subcategoria 4", "sub categoria 4", "sub 4", "sub4"),
        COL_UNIDADE: ("unidade", "unid", "sigla unidade"),
        COL_MODELO: ("modelo",),
        COL_PESO: ("peso", "peso etiqueta", "peso gondola"),
        COL_CODIGO_BARRAS: ("codigo barras", "codigo de barras", "ean", "barras", "cb"),
        COL_PRECO_CUSTO: ("preco custo", "preço custo", "custo", "custo unitario", "custo unitário"),
        COL_PRECO_VENDA: ("preco venda", "preço venda", "venda", "preco de venda", "preço de venda"),
        COL_DEL_ATIVO: ("delivery ativo", "catalogo ativo", "catálogo ativo", "ativo delivery"),
        COL_DEL_TITULO: ("delivery titulo", "delivery título", "titulo delivery", "título delivery"),
        COL_DEL_DESCRICAO: ("delivery descricao", "delivery descrição", "descricao delivery"),
        COL_DEL_ORDEM: ("delivery ordem", "ordem delivery"),
        COL_DEL_DESTAQUE: ("delivery destaque", "destaque delivery"),
        COL_DEL_ESTOQUE_NEG: (
            "delivery estoque neg.",
            "delivery estoque negativo",
            "estoque negativo delivery",
        ),
        COL_DEL_PESO: ("delivery peso", "peso delivery", "peso catalogo", "peso catálogo"),
        COL_DEL_CAT: ("delivery categoria", "categoria delivery", "categoria catalogo"),
        COL_DEL_SUB1: ("delivery sub 1", "delivery sub1", "subcategoria delivery"),
        COL_DEL_SUB2: ("delivery sub 2", "delivery sub2"),
        COL_DEL_SUB3: ("delivery sub 3", "delivery sub3"),
        COL_DEL_SUB4: ("delivery sub 4", "delivery sub4"),
        COL_DEL_EMBALAGENS: ("delivery embalagens", "embalagens delivery", "embalagens catalogo"),
    }
    out: dict[str, str | None] = {}
    for key, keys in aliases.items():
        out[key] = None
        for k in keys:
            if k in norm:
                out[key] = norm[k]
                break
    return out


def _cel_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, bool):
        return "1" if val else "0"
    if isinstance(val, int):
        return str(val)
    if isinstance(val, float):
        if val == int(val) or abs(val - round(val)) < 1e-9:
            return str(int(round(val)))
        s = str(val).strip()
        if "e" in s.lower():
            try:
                return str(int(Decimal(s)))
            except Exception:
                pass
        return s
    s = str(val).strip()
    if s.endswith(".0") and s[:-2].replace("-", "").isdigit():
        return s[:-2]
    return s


def _cel_opt_str(val) -> str | None:
    if val is None:
        return None
    s = _cel_str(val)
    return s if s else None


def _id_produto_planilha_valido(pid: str) -> bool:
    """ID exportado: ObjectId Mongo, Id numérico ERP ou id Postgres Agro (AGRO…)."""
    s = str(pid or "").strip()
    if not s or len(s) > 64:
        return False
    low = s.lower()
    if re.fullmatch(r"[0-9a-f]{24}", low):
        return True
    if s.isdigit() and 1 <= len(s) <= 12:
        return True
    if re.fullmatch(r"AGRO[0-9A-Fa-f]{8,48}", s):
        return True
    if low.startswith("local:") and low[6:].isdigit():
        return True
    return False


def _id_planilha_resumo(pid: str, max_len: int = 28) -> str:
    s = str(pid or "").strip()
    if len(s) <= max_len:
        return s
    return s[: max_len - 1] + "…"


def _parse_decimal_br(val) -> Decimal | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    s = str(val).strip()
    if not s:
        return None
    s = s.replace("R$", "").replace(" ", "").strip()
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _ler_planilha(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    suf = path.suffix.lower()
    if suf == ".csv":
        import csv

        for enc in ("utf-8-sig", "latin-1", "cp1252"):
            try:
                with path.open("r", encoding=enc, newline="") as f:
                    reader = csv.DictReader(f, delimiter=";")
                    if reader.fieldnames and len(reader.fieldnames) == 1:
                        f.seek(0)
                        reader = csv.DictReader(f, delimiter=",")
                    headers = list(reader.fieldnames or [])
                    rows = [dict(r) for r in reader]
                    return headers, rows
            except UnicodeDecodeError:
                continue
        raise ValueError("Não foi possível ler o CSV (encoding).")
    if suf in (".xlsx", ".xls"):
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        headers = [str(c or "").strip() for c in next(it, [])]
        rows = []
        for row in it:
            if not any(row):
                continue
            d: dict[str, Any] = {}
            for i, h in enumerate(headers):
                if h:
                    d[h] = row[i] if i < len(row) else None
            rows.append(d)
        wb.close()
        return headers, rows
    raise ValueError("Use arquivo .csv ou .xlsx.")


def coletar_linhas_export_cadastro(
    *,
    inativos: bool = False,
    categorias: list[str] | None = None,
) -> tuple[list[dict], bool]:
    """Retorna linhas mescladas (Mongo/Agro + overlay) para exportação."""
    from produtos.agro_fonte_config import agro_catalogo_usa_postgres
    from produtos.views import (
        _CADASTRO_LISTA_MONGO_PROJ,
        _aplicar_produto_gestao_overlay_em_dict,
        _overlay_mapa_por_ids_chunked,
        _produto_mongo_para_cadastro_row,
        obter_conexao_mongo,
    )

    rows: list[dict] = []
    truncado = False

    if agro_catalogo_usa_postgres():
        from produtos import catalogo_agro

        pagina = 1
        por_pagina = 500
        while len(rows) < EXPORT_MAX_ROWS:
            chunk, has_more = catalogo_agro.listar_paginado(
                pagina=pagina,
                por_pagina=por_pagina,
                sort_key="nome",
                sort_direction=1,
                inativos=inativos,
            )
            if not chunk:
                break
            pids = [str(r.get("id") or "") for r in chunk]
            ovs = _overlay_mapa_por_ids_chunked(pids)
            for r in chunk:
                _aplicar_produto_gestao_overlay_em_dict(r, ovs.get(str(r.get("id") or "")))
                rows.append(r)
            _enriquecer_rows_delivery_batch(chunk, ovs)
            if not has_more:
                break
            pagina += 1
        truncado = len(rows) >= EXPORT_MAX_ROWS
        rows = _filtrar_rows_categorias(rows[:EXPORT_MAX_ROWS], categorias or [])
        return rows, truncado

    client, db = obter_conexao_mongo()
    if db is None:
        raise ValueError("Mongo indisponível — não foi possível exportar o catálogo.")

    filtro = {} if inativos else {"CadastroInativo": {"$ne": True}}
    cur = (
        db[client.col_p]
        .find(filtro, _CADASTRO_LISTA_MONGO_PROJ)
        .sort("Nome", 1)
        .limit(EXPORT_MAX_ROWS + 1)
    )
    chunk = list(cur)
    truncado = len(chunk) > EXPORT_MAX_ROWS
    chunk = chunk[:EXPORT_MAX_ROWS]
    rows = [_produto_mongo_para_cadastro_row(p) for p in chunk]
    ovs = _overlay_mapa_por_ids_chunked([str(r.get("id") or "") for r in rows])
    for r in rows:
        _aplicar_produto_gestao_overlay_em_dict(r, ovs.get(str(r.get("id") or "")))
    _enriquecer_rows_delivery_batch(rows, ovs)
    rows = _filtrar_rows_categorias(rows, categorias or [])
    return rows, truncado


def enriquecer_rows_ultimos_fornecedores(rows: list[dict], colunas: list[str] | None) -> None:
    """Preenche fornecedor_compra_1..3 se pedidas no Excel ↓ (Entrada NF Agro)."""
    cols = set(colunas or [])
    if not cols.intersection(FORNECEDOR_EXPORT_KEYS):
        return
    if not rows:
        return
    pids = [str(r.get("id") or "").strip() for r in rows if str(r.get("id") or "").strip()]
    if not pids:
        return
    try:
        from produtos.compras_ultimas_compras_util import ultimos_fornecedores_por_produto_ids

        mapa = ultimos_fornecedores_por_produto_ids(pids, n=3)
    except Exception:
        mapa = {}
    for r in rows:
        pid = str(r.get("id") or "").strip()
        nomes = mapa.get(pid) or []
        r[COL_FORN_COMPRA_1] = nomes[0] if len(nomes) > 0 else ""
        r[COL_FORN_COMPRA_2] = nomes[1] if len(nomes) > 1 else ""
        r[COL_FORN_COMPRA_3] = nomes[2] if len(nomes) > 2 else ""


def linha_export_planilha(row: dict) -> dict[str, Any]:
    empty = _delivery_planilha_de_dict({})
    out = {
        COL_ID: str(row.get("id") or ""),
        COL_CODIGO_GM: str(row.get("codigo_nfe") or row.get("codigo") or ""),
        COL_NOME: str(row.get("nome") or ""),
        COL_MARCA: str(row.get("marca") or ""),
        COL_CATEGORIA: str(row.get("categoria") or ""),
        COL_SUBCATEGORIA: str(row.get("subcategoria") or ""),
        COL_SUBCATEGORIA_2: str(row.get("subcategoria_2") or ""),
        COL_SUBCATEGORIA_3: str(row.get("subcategoria_3") or ""),
        COL_SUBCATEGORIA_4: str(row.get("subcategoria_4") or ""),
        COL_UNIDADE: str(row.get("unidade") or ""),
        COL_MODELO: str(row.get("modelo") or ""),
        COL_PESO: str(row.get("peso_etiqueta") or ""),
        COL_CODIGO_BARRAS: str(row.get("codigo_barras") or ""),
        COL_PRECO_CUSTO: float(row.get("preco_custo") or 0),
        COL_PRECO_VENDA: float(row.get("preco_venda") or 0),
    }
    for k in DELIVERY_IMPORT_KEYS:
        if k in row:
            out[k] = str(row.get(k) or "")
        else:
            out[k] = empty.get(k, "")
    for k in FORNECEDOR_EXPORT_KEYS:
        out[k] = str(row.get(k) or "")
    return out


def montar_xlsx_cadastro(rows: list[dict], colunas: list[str] | None = None) -> bytes:
    from openpyxl.styles import Protection

    hdrs = headers_export(colunas)
    wb = Workbook()
    ws = wb.active
    ws.title = "Cadastro"
    hdr_fill = PatternFill("solid", fgColor="DCFCE7")
    hdr_font = Font(bold=True, color="14532D")
    lock_fill = PatternFill("solid", fgColor="F1F5F9")
    for col, (label, key) in enumerate(hdrs, start=1):
        c = ws.cell(row=1, column=col, value=label)
        c.font = hdr_font
        c.fill = hdr_fill
        c.protection = Protection(locked=key in EXPORT_COLS_BLOQUEADAS)
    for ri, src in enumerate(rows, start=2):
        line = linha_export_planilha(src)
        for col, (_, key) in enumerate(hdrs, start=1):
            val = line.get(key)
            cell = ws.cell(row=ri, column=col)
            if key in _COLS_TEXTO_EXCEL:
                cell.value = str(val) if val is not None else ""
                cell.number_format = "@"
            else:
                cell.value = val
                if key in (COL_PRECO_CUSTO, COL_PRECO_VENDA):
                    cell.number_format = "#,##0.00"
            bloqueada = key in EXPORT_COLS_BLOQUEADAS
            cell.protection = Protection(locked=bloqueada)
            if bloqueada:
                cell.fill = lock_fill
    for col in range(1, len(hdrs) + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 16
        key = hdrs[col - 1][1]
        if key in EXPORT_COLS_OCULTAS:
            ws.column_dimensions[letter].hidden = True
            ws.column_dimensions[letter].width = 2
    nome_col = next(
        (get_column_letter(i + 1) for i, (_, key) in enumerate(hdrs) if key == COL_NOME),
        None,
    )
    if nome_col:
        ws.column_dimensions[nome_col].width = 42
    ws.protection.sheet = True
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _estado_atual_produto(pid: str) -> dict | None:
    return _mapa_estado_atual_produtos([pid]).get(str(pid or "").strip()[:64])


def _mapa_estado_atual_produtos(
    pids: list[str],
    on_progress: None | Any = None,
) -> dict[str, dict]:
    """Carrega estado atual (Mongo + overlay) em lote — evita N consultas na prévia."""
    from bson import ObjectId

    from produtos.views import (
        _CADASTRO_LISTA_MONGO_PROJ,
        _aplicar_produto_gestao_overlay_em_dict,
        _mongo_filtro_id_produto_externo,
        _overlay_mapa_por_ids_chunked,
        _produto_mongo_para_cadastro_row,
        _produto_mongo_por_id_externo,
        obter_conexao_mongo,
    )

    uniq = list(dict.fromkeys(str(p or "").strip()[:64] for p in pids if str(p or "").strip()))
    if not uniq:
        return {}

    from produtos.agro_fonte_config import agro_catalogo_usa_postgres

    if agro_catalogo_usa_postgres():
        from produtos import catalogo_agro
        from produtos.models import Produto

        if on_progress:
            on_progress(10, "Catálogo Postgres…")
        ovs = _overlay_mapa_por_ids_chunked(uniq)
        prod_map = {
            str(p.produto_externo_id or "")[:64]: p
            for p in Produto.objects.filter(produto_externo_id__in=uniq)
        }
        out: dict[str, dict] = {}
        for pid in uniq:
            p = prod_map.get(pid) or catalogo_agro.obter_produto_model(pid)
            if p is None:
                continue
            pid_key = str(p.produto_externo_id or pid)[:64]
            row = catalogo_agro.produto_agro_para_row(p, ovs.get(pid_key) or ovs.get(pid))
            out[pid] = row
        _enriquecer_rows_delivery_batch(list(out.values()), ovs)
        return out

    client, db = obter_conexao_mongo()
    if db is None:
        return {}

    doc_map: dict[str, dict] = {}
    chunk_sz = 250
    n_chunks = max(1, (len(uniq) + chunk_sz - 1) // chunk_sz)

    def _registrar_doc(doc: dict, chunk_set: set[str]) -> None:
        id_val = doc.get("Id")
        if id_val is not None:
            sid = str(id_val).strip()[:64]
            if sid in chunk_set:
                doc_map[sid] = doc
                return
        sid = str(doc.get("_id") or "").strip()[:64]
        if sid in chunk_set:
            doc_map[sid] = doc

    for ci, start in enumerate(range(0, len(uniq), chunk_sz)):
        if on_progress:
            on_progress(8 + int(2 * ci / n_chunks), f"Catálogo Mongo… lote {ci + 1}/{n_chunks}")

        chunk_pids = uniq[start : start + chunk_sz]
        chunk_set = set(chunk_pids)
        ors: list[dict] = [{"Id": {"$in": chunk_pids}}]
        int_ids: list[int] = []
        for p in chunk_pids:
            try:
                int_ids.append(int(p))
            except (TypeError, ValueError):
                pass
        if int_ids:
            ors.append({"Id": {"$in": int_ids}})
        oids: list[ObjectId] = []
        for p in chunk_pids:
            try:
                oids.append(ObjectId(p))
            except Exception:
                pass
        if oids:
            ors.append({"_id": {"$in": oids}})

        try:
            for doc in db[client.col_p].find({"$or": ors}, _CADASTRO_LISTA_MONGO_PROJ):
                _registrar_doc(doc, chunk_set)
        except Exception:
            pass

    missing = [p for p in uniq if p not in doc_map]
    if missing:
        for mi in range(0, len(missing), 40):
            sub = missing[mi : mi + 40]
            sub_set = set(sub)
            ors_fb: list[dict] = []
            for pid in sub:
                filt = _mongo_filtro_id_produto_externo(pid)
                ors_fb.extend(filt.get("$or") or [filt])
            try:
                for doc in db[client.col_p].find({"$or": ors_fb}, _CADASTRO_LISTA_MONGO_PROJ):
                    _registrar_doc(doc, sub_set)
            except Exception:
                for pid in sub:
                    if pid in doc_map:
                        continue
                    doc = _produto_mongo_por_id_externo(db, client, pid)
                    if doc:
                        doc_map[pid] = doc

    if on_progress:
        on_progress(10, "Mesclando overlay Agro…")

    ovs = _overlay_mapa_por_ids_chunked(uniq)
    out: dict[str, dict] = {}
    for pid in uniq:
        doc = doc_map.get(pid)
        if not doc:
            continue
        row = _produto_mongo_para_cadastro_row(doc)
        _aplicar_produto_gestao_overlay_em_dict(row, ovs.get(pid))
        out[pid] = row
    _enriquecer_rows_delivery_batch(list(out.values()), ovs)
    return out


def _patch_da_linha(raw: dict, colmap: dict[str, str | None]) -> dict[str, Any]:
    patch: dict[str, Any] = {}

    def txt(key: str, mx: int) -> None:
        hdr = colmap.get(key)
        if not hdr or hdr not in raw:
            return
        v = _cel_opt_str(raw.get(hdr))
        if v is not None:
            patch[key] = v[:mx]

    def dec(key: str) -> None:
        hdr = colmap.get(key)
        if not hdr or hdr not in raw:
            return
        v = raw.get(hdr)
        if v is None or (isinstance(v, str) and not str(v).strip()):
            return
        d = _parse_decimal_br(v)
        if d is None:
            patch[f"__erro_{key}"] = f"Valor inválido em «{hdr}»."
        else:
            patch[key] = d

    txt(COL_CODIGO_GM, 64)
    txt(COL_NOME, 300)
    txt(COL_MARCA, 120)
    txt(COL_MODELO, 200)
    txt(COL_UNIDADE, 20)
    txt(COL_PESO, 40)
    txt(COL_CATEGORIA, 200)
    txt(COL_SUBCATEGORIA, 200)
    txt(COL_SUBCATEGORIA_2, 200)
    txt(COL_SUBCATEGORIA_3, 200)
    txt(COL_SUBCATEGORIA_4, 200)
    txt(COL_CODIGO_BARRAS, 80)
    dec(COL_PRECO_CUSTO)
    dec(COL_PRECO_VENDA)
    txt(COL_DEL_ATIVO, 10)
    txt(COL_DEL_TITULO, 200)
    txt(COL_DEL_DESCRICAO, 2000)
    txt(COL_DEL_ORDEM, 10)
    txt(COL_DEL_DESTAQUE, 10)
    txt(COL_DEL_ESTOQUE_NEG, 10)
    txt(COL_DEL_PESO, 40)
    txt(COL_DEL_CAT, 80)
    txt(COL_DEL_SUB1, 80)
    txt(COL_DEL_SUB2, 80)
    txt(COL_DEL_SUB3, 80)
    txt(COL_DEL_SUB4, 80)
    txt(COL_DEL_EMBALAGENS, 500)
    return patch


def _merged_row(atual: dict, patch: dict) -> dict:
    out = dict(atual)
    for k, v in patch.items():
        if k.startswith("__"):
            continue
        if k in (COL_PRECO_CUSTO, COL_PRECO_VENDA):
            out[k] = float(v)
        elif k == COL_CODIGO_GM:
            out["codigo_gm"] = v
            out["codigo_nfe"] = v
        else:
            out[k] = v
    return out


def _validar_merged(row: dict) -> str | None:
    if not str(row.get("nome") or "").strip():
        return "Nome obrigatório."
    if not str(row.get("marca") or "").strip():
        return "Marca obrigatória."
    if not str(row.get("categoria") or "").strip():
        return "Categoria obrigatória."
    if not str(row.get("codigo_barras") or "").strip():
        return "Código de barras obrigatório."
    try:
        Decimal(str(row.get("preco_venda") or ""))
    except Exception:
        return "Preço de venda inválido."
    try:
        Decimal(str(row.get("preco_custo") or ""))
    except Exception:
        return "Preço de custo inválido."
    return None


def _tem_alteracao(atual: dict, patch: dict) -> bool:
    for k in IMPORT_KEYS:
        if k not in patch:
            continue
        if k in (COL_PRECO_CUSTO, COL_PRECO_VENDA):
            if round(float(atual.get(k) or 0), 2) != round(float(patch[k]), 2):
                return True
        elif _valor_atual_campo_import(atual, k) != str(patch[k] or "").strip():
            return True
    return False


def preview_importacao_cadastro(
    path: Path,
    on_progress: None | Any = None,
) -> dict[str, Any]:
    headers, rows_raw = _ler_planilha(path)
    if on_progress:
        on_progress(0, f"total:{len(rows_raw)}")
        on_progress(2, f"Planilha com {len(rows_raw)} linha(s) — lendo…")
    colmap = _map_headers(headers)
    if not colmap.get(COL_ID):
        raise ValueError("Coluna «ID» não encontrada na planilha.")

    hdr_id = colmap[COL_ID]
    rows_slice = rows_raw[:IMPORT_MAX_ROWS]
    vistos: set[str] = set()
    alteracoes: list[dict] = []
    ignoradas: list[dict] = []
    erros: list[dict] = []
    pendentes: list[dict] = []

    for i, raw in enumerate(rows_slice, start=2):
        pid = _cel_str(raw.get(hdr_id or ""))[:64]
        if not pid:
            continue
        if not _id_produto_planilha_valido(pid):
            erros.append(
                {
                    "linha": i,
                    "id": _id_planilha_resumo(pid),
                    "erro": "ID inválido (texto ou coluna errada). Apague a linha ou restaure o ID da exportação.",
                }
            )
            continue
        if pid in vistos:
            erros.append({"linha": i, "id": pid, "erro": "ID duplicado na planilha."})
            continue
        vistos.add(pid)

        patch = _patch_da_linha(raw, colmap)
        err_fields = [v for k, v in patch.items() if k.startswith("__erro_")]
        if err_fields:
            erros.append({"linha": i, "id": pid, "erro": err_fields[0]})
            continue
        if not any(k in patch for k in IMPORT_KEYS):
            ignoradas.append({"linha": i, "id": pid, "motivo": "Nenhum campo alterado (células vazias)."})
            continue
        pendentes.append({"linha": i, "id": pid, "patch": patch})

    if on_progress:
        on_progress(8, f"Conferindo {len(pendentes)} linha(s) preenchida(s)…")

    try:
        facetas = carregar_facetas_planilha()
    except Exception:
        facetas = {k: [] for k in _COLS_FACETA}
    canonicos = {k: _mapa_canonico_faceta(facetas.get(k) or []) for k in _COLS_FACETA}

    mapa = _mapa_estado_atual_produtos([p["id"] for p in pendentes], on_progress=on_progress)
    total_pend = len(pendentes)
    step = max(1, total_pend // 40)
    eventos_faceta: list[dict] = []

    for idx, item in enumerate(pendentes):
        pid = item["id"]
        patch = item["patch"]
        i = item["linha"]

        if on_progress and (idx == 0 or idx % step == 0 or idx == total_pend - 1):
            pct = 10 + int(85 * idx / max(1, total_pend))
            on_progress(pct, f"Analisando linha {i}… ({idx + 1}/{total_pend})")

        atual = mapa.get(pid)
        if not atual:
            erros.append({"linha": i, "id": pid, "erro": "Produto não encontrado no catálogo."})
            continue

        patch_res, evs, _errs_fac = _resolver_facetas_no_patch(
            patch, facetas, canonicos, permitir_novos=True, linha=i
        )
        eventos_faceta.extend(evs)

        if not _tem_alteracao(atual, patch_res):
            if any(e.get("acao") == "novo" for e in evs):
                ignoradas.append({"linha": i, "id": pid, "motivo": "Valores iguais ao cadastro atual."})
                continue
            if not _tem_alteracao(atual, patch):
                ignoradas.append({"linha": i, "id": pid, "motivo": "Valores iguais ao cadastro atual."})
                continue
            ignoradas.append(
                {
                    "linha": i,
                    "id": pid,
                    "motivo": "Após corrigir typo, valores iguais ao cadastro.",
                }
            )
            continue

        verr_del = _validar_patch_delivery(atual, patch_res)
        if verr_del:
            erros.append({"linha": i, "id": pid, "erro": verr_del})
            continue

        merged = _merged_row(atual, patch_res)
        vmsg = _validar_merged(merged)
        if vmsg:
            erros.append({"linha": i, "id": pid, "erro": vmsg})
            continue

        novos_na_linha = [e for e in evs if e.get("acao") == "novo"]
        if novos_na_linha:
            for nv in novos_na_linha:
                erros.append(
                    {
                        "linha": i,
                        "id": pid,
                        "erro": (
                            f"{nv.get('rotulo')} «{nv.get('valor')}» ainda não cadastrado. "
                            "Marque «Permitir criar novos» para gravar, ou use a lista / sugestão."
                        ),
                        "tipo": "valor_novo",
                        "campo": nv.get("campo"),
                        "valor": nv.get("valor"),
                    }
                )

        campos = []
        for k in IMPORT_KEYS:
            if k not in patch_res:
                continue
            item_c = {
                "campo": k,
                "de": atual.get(k),
                "para": patch_res[k],
            }
            for ev in evs:
                if ev.get("campo") == k and ev.get("acao") == "corrigir":
                    item_c["sugestao"] = ev.get("sugestao")
                    item_c["de_planilha"] = ev.get("valor")
                    item_c["nota"] = f"typo → {ev.get('sugestao')}"
                elif ev.get("campo") == k and ev.get("acao") == "novo":
                    item_c["nota"] = "valor novo"
            campos.append(item_c)
        if not campos:
            continue
        alteracoes.append(
            {
                "linha": i,
                "id": pid,
                "nome": merged.get("nome") or atual.get("nome") or "",
                "campos": campos,
                "tem_valor_novo": bool(novos_na_linha),
            }
        )

    if on_progress:
        on_progress(100, "Concluído")

    if len(rows_raw) > IMPORT_MAX_ROWS:
        erros.append(
            {
                "linha": None,
                "id": "",
                "erro": f"Planilha truncada: só as primeiras {IMPORT_MAX_ROWS} linhas foram analisadas.",
            }
        )

    valores_novos, correcoes = _resumir_eventos_faceta(eventos_faceta)
    n_ok = sum(1 for a in alteracoes if not a.get("tem_valor_novo"))
    n_bloqueadas_novo = sum(1 for a in alteracoes if a.get("tem_valor_novo"))

    return {
        "total_linhas": len(rows_raw),
        "alteracoes": alteracoes[:500],
        "n_alteracoes": len(alteracoes),
        "n_alteracoes_ok": n_ok,
        "n_bloqueadas_valor_novo": n_bloqueadas_novo,
        "ignoradas": ignoradas[:80],
        "n_ignoradas": len(ignoradas),
        "erros": erros[:120],
        "n_erros": len(erros),
        "valores_novos": valores_novos[:80],
        "n_valores_novos": len(valores_novos),
        "correcoes_sugeridas": correcoes[:80],
        "n_correcoes_sugeridas": len(correcoes),
        "permitir_novos_padrao": False,
    }


def _snapshot_antes_import(db, client, pid: str, patch: dict, nome: str = "") -> dict:
    """Estado overlay + Mongo antes de gravar — usado para desfazer."""
    from produtos.views import _produto_mongo_por_id_externo

    pid = str(pid or "").strip()[:64]
    ov = ProdutoGestaoOverlayAgro.objects.filter(produto_externo_id=pid).first()
    overlay: dict[str, Any] = {}
    for k in OVERLAY_IMPORT_KEYS:
        overlay[k] = _ler_overlay_import_campo(ov, k)
    from produtos.catalogo_delivery_util import delivery_de_extras

    if any(k in patch for k in DELIVERY_IMPORT_KEYS):
        overlay["_delivery_blob"] = delivery_de_extras(_extras_dict(ov) if ov else {})

    mongo: dict[str, Any] = {}
    if db is not None:
        doc = _produto_mongo_por_id_externo(db, client, pid)
        if doc:
            for mk in ("PrecoCusto", "ValorCusto", "ValorVenda", "PrecoVenda"):
                if mk in doc and doc[mk] is not None:
                    mongo[mk] = float(doc[mk])

    produto_pg: dict[str, Any] = {}
    from produtos.agro_fonte_config import agro_catalogo_usa_postgres

    if agro_catalogo_usa_postgres():
        from produtos import catalogo_agro

        p_pg = catalogo_agro.obter_produto_model(pid)
        if p_pg is not None:
            produto_pg = _snapshot_produto_pg(p_pg)

    campos = [k for k in IMPORT_KEYS if k in patch]
    para: dict[str, Any] = {}
    for k in campos:
        v = patch[k]
        if k in (COL_PRECO_CUSTO, COL_PRECO_VENDA):
            para[k] = float(v)
        else:
            para[k] = v

    return {
        "id": pid,
        "nome": str(nome or "")[:300],
        "campos_alterados": campos,
        "overlay_existia": ov is not None,
        "overlay": overlay,
        "mongo": mongo,
        "produto_pg": produto_pg,
        "para": para,
    }


def _overlay_import_esta_vazio(ov: ProdutoGestaoOverlayAgro) -> bool:
    for k in OVERLAY_IMPORT_KEYS:
        if k in DELIVERY_IMPORT_KEYS:
            continue
        v = _ler_overlay_import_campo(ov, k)
        if k == COL_PRECO_VENDA:
            if v is not None:
                return False
        elif str(v or "").strip():
            return False
    ex = _extras_dict(ov)
    if isinstance(ex.get("delivery"), dict) and ex.get("delivery"):
        return False
    return True


def _reverter_item_import(item: dict, db, client, user) -> None:
    from produtos.views import _mongo_filtro_id_produto_externo

    pid = str(item.get("id") or "").strip()[:64]
    if not pid:
        return
    patch_keys = list(item.get("campos_alterados") or item.get("para", {}).keys())
    overlay_antes = item.get("overlay") or {}
    overlay_existia = bool(item.get("overlay_existia"))

    ov = ProdutoGestaoOverlayAgro.objects.filter(produto_externo_id=pid).first()

    if overlay_existia:
        if not ov:
            ov = ProdutoGestaoOverlayAgro.objects.create(
                produto_externo_id=pid,
                usuario=user if user and user.is_authenticated else None,
            )
        for k in patch_keys:
            if k not in OVERLAY_IMPORT_KEYS:
                continue
            if k in DELIVERY_IMPORT_KEYS:
                continue
            _gravar_overlay_import_campo(ov, k, overlay_antes.get(k))
        if any(k in DELIVERY_IMPORT_KEYS for k in patch_keys):
            from produtos.catalogo_delivery_util import normalizar_delivery

            ex = _extras_dict(ov)
            blob = overlay_antes.get("_delivery_blob")
            if isinstance(blob, dict) and blob:
                ex["delivery"] = normalizar_delivery(blob, processar_imagem=False)
            else:
                ex.pop("delivery", None)
            ov.cadastro_extras = ex
        ov.save()
    elif ov:
        for k in patch_keys:
            if k not in OVERLAY_IMPORT_KEYS:
                continue
            if k in DELIVERY_IMPORT_KEYS:
                continue
            _gravar_overlay_import_campo(ov, k, None if k == COL_PRECO_VENDA else "")
        if any(k in DELIVERY_IMPORT_KEYS for k in patch_keys):
            ex = _extras_dict(ov)
            ex.pop("delivery", None)
            ov.cadastro_extras = ex
        if _overlay_import_esta_vazio(ov):
            ov.delete()
        else:
            ov.save()

    mongo_antes = item.get("mongo") or {}
    if db is not None and mongo_antes:
        mongo_set: dict[str, float] = {}
        if COL_PRECO_CUSTO in patch_keys:
            for mk in ("PrecoCusto", "ValorCusto"):
                if mk in mongo_antes:
                    mongo_set[mk] = float(mongo_antes[mk])
        if COL_PRECO_VENDA in patch_keys:
            for mk in ("ValorVenda", "PrecoVenda"):
                if mk in mongo_antes:
                    mongo_set[mk] = float(mongo_antes[mk])
        if mongo_set:
            from produtos.agro_mongo_guard import agro_mongo_escrita_bloqueada

            if not agro_mongo_escrita_bloqueada():
                db[client.col_p].update_one(_mongo_filtro_id_produto_externo(pid), {"$set": mongo_set})

    from produtos.agro_fonte_config import agro_catalogo_usa_postgres

    if agro_catalogo_usa_postgres():
        from produtos import catalogo_agro

        pg_antes = item.get("produto_pg") or {}
        p = catalogo_agro.obter_produto_model(pid)
        if p is not None and pg_antes:
            if _restaurar_produto_pg(p, pg_antes, patch_keys):
                p.save()


def _historico_import_resumo_item(item: dict) -> dict:
    campos = item.get("campos_alterados") or []
    detalhes = []
    para = item.get("para") or {}
    de_map = item.get("de_merged") or item.get("overlay") or {}
    for k in campos[:6]:
        detalhes.append({"campo": k, "de": de_map.get(k), "para": para.get(k)})
    return {
        "id": item.get("id"),
        "nome": item.get("nome") or "",
        "campos": campos,
        "detalhes": detalhes,
    }


def listar_historico_import_cadastro(*, limite: int = HISTORICO_IMPORT_LISTA_LIMITE) -> list[dict]:
    out: list[dict] = []
    qs = CadastroPlanilhaImportHistoricoAgro.objects.select_related("usuario", "revertido_por")
    if hasattr(CadastroPlanilhaImportHistoricoAgro, "tipo"):
        qs = qs.filter(tipo=CadastroPlanilhaImportHistoricoAgro.Tipo.CADASTRO)
    qs = qs.order_by("-criado_em")[:limite]
    for h in qs:
        items = (h.backup or {}).get("items") or []
        out.append(
            {
                "id": h.pk,
                "criado_em": h.criado_em.isoformat() if h.criado_em else "",
                "usuario": (h.usuario.get_username() if h.usuario else "") or "",
                "nome_arquivo": h.nome_arquivo or "",
                "n_produtos": h.n_produtos,
                "n_campos": h.n_campos,
                "status": h.status,
                "pode_reverter": h.status == CadastroPlanilhaImportHistoricoAgro.Status.APLICADO,
                "revertido_em": h.revertido_em.isoformat() if h.revertido_em else "",
                "revertido_por": (h.revertido_por.get_username() if h.revertido_por else "") or "",
                "resumo": [_historico_import_resumo_item(it) for it in items[:12]],
            }
        )
    return out


def reverter_importacao_cadastro(historico_id: int, user) -> dict[str, Any]:
    from django.utils import timezone

    hist = CadastroPlanilhaImportHistoricoAgro.objects.filter(pk=historico_id).first()
    if not hist:
        raise ValueError("Histórico não encontrado.")
    if getattr(hist, "tipo", "cadastro") == "estoque" or (hist.backup or {}).get("tipo") == "estoque":
        raise ValueError("Use o histórico de Excel estoque para desfazer esta importação.")
    if hist.status != CadastroPlanilhaImportHistoricoAgro.Status.APLICADO:
        raise ValueError("Esta importação já foi desfeita.")

    items = (hist.backup or {}).get("items") or []
    if not items:
        raise ValueError("Backup vazio — não é possível desfazer.")

    from produtos.agro_fonte_config import agro_catalogo_usa_postgres
    from produtos.views import obter_conexao_mongo

    use_pg = agro_catalogo_usa_postgres()
    client, db = (None, None)
    if not use_pg:
        client, db = obter_conexao_mongo()
        if db is None:
            raise ValueError("Mongo indisponível.")

    with transaction.atomic():
        for item in items:
            _reverter_item_import(item, db, client, user)
        hist.status = CadastroPlanilhaImportHistoricoAgro.Status.REVERTIDO
        hist.revertido_em = timezone.now()
        hist.revertido_por = user if user and user.is_authenticated else None
        hist.save(update_fields=["status", "revertido_em", "revertido_por"])

    _invalidar_cache_catalogo_pdv()
    return {
        "historico_id": hist.pk,
        "revertidos": len(items),
        "status": hist.status,
    }


def _invalidar_cache_catalogo_pdv() -> None:
    from produtos.views import CATALOGO_PDV_CACHE_ENTRY_KEY, CATALOGO_PDV_CACHE_PREV_ENTRY_KEY

    try:
        cur_cat = cache.get(CATALOGO_PDV_CACHE_ENTRY_KEY)
        if isinstance(cur_cat, dict) and cur_cat.get("version"):
            cache.set(CATALOGO_PDV_CACHE_PREV_ENTRY_KEY, cur_cat, timeout=86400 * 3)
        cache.delete(CATALOGO_PDV_CACHE_ENTRY_KEY)
    except Exception:
        pass


def _gravar_patch_produto(db, client, pid: str, patch: dict, user) -> None:
    from produtos.cadastro_alteracao_historico_util import (
        registrar_diffs_cadastro,
        snapshot_overlay,
    )
    from produtos.models import ProdutoCadastroAlteracaoAgro
    from produtos.views import _mongo_filtro_id_produto_externo

    ov, _ = ProdutoGestaoOverlayAgro.objects.get_or_create(
        produto_externo_id=pid[:64],
        defaults={"usuario": user if user and user.is_authenticated else None},
    )
    antes = snapshot_overlay(ov)
    for k in OVERLAY_IMPORT_KEYS:
        if k in patch and k not in DELIVERY_IMPORT_KEYS:
            _gravar_overlay_import_campo(ov, k, patch[k])
    err_del = _aplicar_patch_delivery(ov, patch)
    if err_del:
        raise ValueError(err_del)
    ex = _extras_dict(ov)
    if COL_PRECO_CUSTO in patch:
        ex["preco_custo_overlay"] = float(patch[COL_PRECO_CUSTO])
    ov.cadastro_extras = ex
    depois = snapshot_overlay(ov)
    try:
        registrar_diffs_cadastro(
            produto_id=pid,
            antes=antes,
            depois=depois,
            usuario=user if user and getattr(user, "is_authenticated", False) else None,
            origem=ProdutoCadastroAlteracaoAgro.Origem.PLANILHA,
        )
    except Exception:
        pass
    ov.save()

    custo_payload = patch.get(COL_PRECO_CUSTO) if COL_PRECO_CUSTO in patch else None

    from produtos.agro_fonte_config import agro_catalogo_usa_postgres

    if agro_catalogo_usa_postgres():
        _aplicar_patch_no_produto_pg(pid, ov, patch, custo_payload=custo_payload)

    mongo_set: dict[str, float] = {}
    if COL_PRECO_CUSTO in patch:
        cfloat = float(patch[COL_PRECO_CUSTO])
        mongo_set["PrecoCusto"] = cfloat
        mongo_set["ValorCusto"] = cfloat
    if COL_PRECO_VENDA in patch:
        pvfloat = float(patch[COL_PRECO_VENDA])
        mongo_set["ValorVenda"] = pvfloat
        mongo_set["PrecoVenda"] = pvfloat
    if mongo_set and db is not None:
        from produtos.agro_mongo_guard import agro_mongo_escrita_bloqueada

        if not agro_mongo_escrita_bloqueada():
            db[client.col_p].update_one(_mongo_filtro_id_produto_externo(pid), {"$set": mongo_set})

    if COL_PRECO_CUSTO in patch and custo_payload is not None:
        try:
            from produtos.custo_familia_util import propagar_custo_familia_de_pai

            propagar_custo_familia_de_pai(
                pid, custo_payload, origem="planilha", usuario=user
            )
        except Exception:
            pass


def aplicar_importacao_cadastro(
    path: Path,
    user,
    *,
    nome_arquivo: str = "",
    permitir_novos: bool = False,
    on_progress: None | Any = None,
) -> dict[str, Any]:
    from produtos.agro_fonte_config import agro_catalogo_usa_postgres
    from produtos.views import obter_conexao_mongo

    use_pg = agro_catalogo_usa_postgres()
    client, db = (None, None)
    if not use_pg:
        client, db = obter_conexao_mongo()
        if db is None:
            raise ValueError("Mongo indisponível.")

    headers, rows_raw = _ler_planilha(path)
    colmap = _map_headers(headers)
    hdr_id = colmap.get(COL_ID)
    if not hdr_id:
        raise ValueError("Coluna «ID» não encontrada.")

    if on_progress:
        on_progress(2, f"total:{len(rows_raw)}")
        on_progress(4, f"Lendo {len(rows_raw)} linha(s)…")

    try:
        facetas = carregar_facetas_planilha()
    except Exception:
        facetas = {k: [] for k in _COLS_FACETA}
    canonicos = {k: _mapa_canonico_faceta(facetas.get(k) or []) for k in _COLS_FACETA}

    candidatos: list[dict] = []
    vistos: set[str] = set()
    for i, raw in enumerate(rows_raw[:IMPORT_MAX_ROWS], start=2):
        pid = _cel_str(raw.get(hdr_id or ""))[:64]
        if not pid or not _id_produto_planilha_valido(pid) or pid in vistos:
            continue
        vistos.add(pid)
        patch = _patch_da_linha(raw, colmap)
        if patch.get("__erro_preco_custo") or patch.get("__erro_preco_venda"):
            continue
        if not any(k in patch for k in IMPORT_KEYS):
            continue
        candidatos.append({"linha": i, "id": pid, "patch": patch})

    if on_progress:
        on_progress(8, f"Conferindo {len(candidatos)} linha(s) com dados…")

    mapa = _mapa_estado_atual_produtos([c["id"] for c in candidatos], on_progress=on_progress)

    fila: list[dict] = []
    bloqueados_novo = 0
    for item in candidatos:
        pid = item["id"]
        patch = item["patch"]
        i = item["linha"]
        atual = mapa.get(pid)
        if not atual:
            continue
        patch_res, _evs, errs_fac = _resolver_facetas_no_patch(
            patch, facetas, canonicos, permitir_novos=permitir_novos, linha=i
        )
        if errs_fac:
            bloqueados_novo += 1
            continue
        if not _tem_alteracao(atual, patch_res):
            continue
        verr_del = _validar_patch_delivery(atual, patch_res)
        if verr_del:
            continue
        merged = _merged_row(atual, patch_res)
        vmsg = _validar_merged(merged)
        if vmsg:
            continue
        fila.append({"linha": i, "id": pid, "patch": patch_res, "atual": atual})

    if not fila:
        if bloqueados_novo:
            raise ValueError(
                f"{bloqueados_novo} linha(s) com marca/categoria/sub nova. "
                "Marque «Permitir criar novos» ou use nomes da lista / correção de typo."
            )
        raise ValueError("Nenhuma alteração válida para gravar — confira a prévia.")

    if on_progress:
        on_progress(92, f"Gravando {len(fila)} produto(s)…")

    ok = 0
    falhas: list[dict] = []
    backups: list[dict] = []
    n_campos_total = 0
    total_fila = len(fila)

    for idx, item in enumerate(fila):
        pid = item["id"]
        patch = item["patch"]
        i = item["linha"]
        atual = item["atual"]
        if on_progress and (idx == 0 or idx == total_fila - 1 or idx % max(1, total_fila // 20) == 0):
            pct = 92 + int(7 * idx / max(1, total_fila))
            on_progress(pct, f"Gravando produto {idx + 1}/{total_fila}…")

        snap = _snapshot_antes_import(
            db,
            client,
            pid,
            patch,
            nome=str(atual.get("nome") or ""),
        )
        snap["de_merged"] = {
            k: atual.get(k) for k in snap.get("campos_alterados") or [] if k in patch
        }
        try:
            with transaction.atomic():
                _gravar_patch_produto(db, client, pid, patch, user)
            backups.append(snap)
            n_campos_total += len(snap.get("campos_alterados") or [])
            ok += 1
        except Exception as exc:
            falhas.append({"linha": i, "id": pid, "erro": str(exc) or "Falha ao gravar."})

    historico_id = None
    if backups:
        hist = CadastroPlanilhaImportHistoricoAgro.objects.create(
            usuario=user if user and user.is_authenticated else None,
            nome_arquivo=str(nome_arquivo or "")[:255],
            n_produtos=len(backups),
            n_campos=n_campos_total,
            backup={"items": backups, "permitir_novos": bool(permitir_novos)},
            **(
                {"tipo": CadastroPlanilhaImportHistoricoAgro.Tipo.CADASTRO}
                if hasattr(CadastroPlanilhaImportHistoricoAgro, "tipo")
                else {}
            ),
        )
        historico_id = hist.pk

    if ok:
        _invalidar_cache_catalogo_pdv()

    if on_progress:
        on_progress(100, "Concluído")

    return {
        "gravados": ok,
        "falhas": falhas[:80],
        "n_falhas": len(falhas),
        "historico_id": historico_id,
        "n_alteracoes": ok,
        "n_bloqueados_valor_novo": bloqueados_novo,
        "permitir_novos": bool(permitir_novos),
    }

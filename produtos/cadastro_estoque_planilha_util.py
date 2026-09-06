"""Excel só de estoque no Cadastro ERP — export filtrado + import com prévia.

Colunas editáveis: Saldo Centro/Vila (absoluto) e Ajuste Centro/Vila (+/−).
Última alteração (data/quem) é só informativa — ignorada na importação.
"""

from __future__ import annotations

import re
import unicodedata
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any, Callable

from django.core.cache import cache
from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Protection
from openpyxl.utils import get_column_letter

from estoque.models import AjusteRapidoEstoque, OrigemAjusteEstoque
from produtos.models import CadastroPlanilhaImportHistoricoAgro

EXPORT_MAX_ROWS = 8000
IMPORT_MAX_ROWS = 2500
HISTORICO_LISTA_LIMITE = 30
TIPO_HISTORICO = "estoque"

COL_ID = "id"
COL_CODIGO_GM = "codigo_gm"
COL_NOME = "nome"
COL_CODIGO_BARRAS = "codigo_barras"
COL_SALDO_CENTRO = "saldo_centro"
COL_AJUSTE_CENTRO = "ajuste_centro"
COL_ULT_DATA_CENTRO = "ult_data_centro"
COL_ULT_QUEM_CENTRO = "ult_quem_centro"
COL_SALDO_VILA = "saldo_vila"
COL_AJUSTE_VILA = "ajuste_vila"
COL_ULT_DATA_VILA = "ult_data_vila"
COL_ULT_QUEM_VILA = "ult_quem_vila"

EXPORT_HEADERS: list[tuple[str, str]] = [
    ("ID", COL_ID),
    ("Código GM", COL_CODIGO_GM),
    ("Nome", COL_NOME),
    ("Código barras", COL_CODIGO_BARRAS),
    ("Saldo Centro", COL_SALDO_CENTRO),
    ("Ajuste Centro (+/-)", COL_AJUSTE_CENTRO),
    ("Últ. alt. Centro", COL_ULT_DATA_CENTRO),
    ("Quem alt. Centro", COL_ULT_QUEM_CENTRO),
    ("Saldo Vila", COL_SALDO_VILA),
    ("Ajuste Vila (+/-)", COL_AJUSTE_VILA),
    ("Últ. alt. Vila", COL_ULT_DATA_VILA),
    ("Quem alt. Vila", COL_ULT_QUEM_VILA),
]

EXPORT_COLS_BLOQUEADAS = frozenset(
    {
        COL_ID,
        COL_CODIGO_GM,
        COL_NOME,
        COL_CODIGO_BARRAS,
        COL_ULT_DATA_CENTRO,
        COL_ULT_QUEM_CENTRO,
        COL_ULT_DATA_VILA,
        COL_ULT_QUEM_VILA,
    }
)
EXPORT_COLS_OCULTAS = frozenset({COL_ID})
IMPORT_EDIT_KEYS = frozenset(
    {COL_SALDO_CENTRO, COL_AJUSTE_CENTRO, COL_SALDO_VILA, COL_AJUSTE_VILA}
)

ProgressCb = Callable[[int, str], None] | None


def _norm_header(h: str) -> str:
    s = unicodedata.normalize("NFD", str(h or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.lower().strip()
    s = re.sub(r"\s+", " ", s)
    return s


def _map_headers(headers: list[str]) -> dict[str, str | None]:
    norm = {_norm_header(h): h for h in headers if str(h or "").strip()}
    aliases: dict[str, tuple[str, ...]] = {
        COL_ID: ("id", "produto id", "produto_id"),
        COL_CODIGO_GM: ("codigo gm", "codigo_gm", "codigo nfe", "gm"),
        COL_NOME: ("nome", "produto", "descricao produto"),
        COL_CODIGO_BARRAS: ("codigo barras", "codigo de barras", "ean", "barras", "cb"),
        COL_SALDO_CENTRO: (
            "saldo centro",
            "estoque centro",
            "saldo c",
            "centro",
            "qtd centro",
        ),
        COL_AJUSTE_CENTRO: (
            "ajuste centro (+/-)",
            "ajuste centro",
            "ajuste +/- centro",
            "delta centro",
            "diferenca centro",
        ),
        COL_ULT_DATA_CENTRO: (
            "ult. alt. centro",
            "ult alt centro",
            "ultima alt centro",
            "ultima alteracao centro",
            "data alt centro",
        ),
        COL_ULT_QUEM_CENTRO: (
            "quem alt. centro",
            "quem alt centro",
            "usuario alt centro",
            "operador centro",
        ),
        COL_SALDO_VILA: (
            "saldo vila",
            "estoque vila",
            "saldo v",
            "vila",
            "vila elias",
            "qtd vila",
        ),
        COL_AJUSTE_VILA: (
            "ajuste vila (+/-)",
            "ajuste vila",
            "ajuste +/- vila",
            "delta vila",
            "diferenca vila",
        ),
        COL_ULT_DATA_VILA: (
            "ult. alt. vila",
            "ult alt vila",
            "ultima alt vila",
            "ultima alteracao vila",
            "data alt vila",
        ),
        COL_ULT_QUEM_VILA: (
            "quem alt. vila",
            "quem alt vila",
            "usuario alt vila",
            "operador vila",
        ),
    }
    out: dict[str, str | None] = {}
    for key, keys in aliases.items():
        out[key] = None
        for k in keys:
            if k in norm:
                out[key] = norm[k]
                break
    return out


def _cel_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, bool):
        return "1" if val else "0"
    if isinstance(val, int):
        return str(val)
    if isinstance(val, float):
        if val == int(val) or abs(val - round(val)) < 1e-9:
            return str(int(round(val)))
        return str(val).strip()
    s = str(val).strip()
    if s.endswith(".0") and s[:-2].replace("-", "").isdigit():
        return s[:-2]
    return s


def _id_produto_planilha_valido(pid: str) -> bool:
    """ObjectId Mongo, Id ERP numérico ou id Postgres Agro (AGRO…)."""
    s = str(pid or "").strip()
    if not s or len(s) > 64:
        return False
    low = s.lower()
    if re.fullmatch(r"[0-9a-f]{24}", low):
        return True
    if s.isdigit() and 1 <= len(s) <= 12:
        return True
    if re.fullmatch(r"AGRO[0-9A-Fa-f]{8,48}", s):
        return True
    if low.startswith("local:") and low[6:].isdigit():
        return True
    return False


def _id_planilha_resumo(pid: str, max_len: int = 28) -> str:
    s = str(pid or "").strip()
    if len(s) <= max_len:
        return s
    return s[: max_len - 1] + "…"


def _parse_decimal_br(val) -> Decimal | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    s = str(val).strip()
    if not s:
        return None
    s = s.replace("R$", "").replace(" ", "").strip()
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _round3(d: Decimal) -> Decimal:
    return d.quantize(Decimal("0.001"))


def _quase_igual(a: Decimal, b: Decimal) -> bool:
    return abs(_round3(a) - _round3(b)) <= Decimal("0.001")


def _ler_planilha(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    suf = path.suffix.lower()
    if suf == ".csv":
        import csv

        for enc in ("utf-8-sig", "latin-1", "cp1252"):
            try:
                with path.open("r", encoding=enc, newline="") as f:
                    reader = csv.DictReader(f, delimiter=";")
                    if reader.fieldnames and len(reader.fieldnames) == 1:
                        f.seek(0)
                        reader = csv.DictReader(f, delimiter=",")
                    headers = list(reader.fieldnames or [])
                    rows = [dict(r) for r in reader]
                    return headers, rows
            except UnicodeDecodeError:
                continue
        raise ValueError("Não foi possível ler o CSV (encoding).")
    if suf in (".xlsx", ".xls"):
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        headers = [str(c or "").strip() for c in next(it, [])]
        rows = []
        for row in it:
            if not any(row):
                continue
            d: dict[str, Any] = {}
            for i, h in enumerate(headers):
                if h:
                    d[h] = row[i] if i < len(row) else None
            rows.append(d)
        wb.close()
        return headers, rows
    raise ValueError("Use arquivo .csv ou .xlsx.")


def _fmt_dt_ajuste(dt) -> str:
    if not dt:
        return ""
    try:
        local = timezone.localtime(dt) if timezone.is_aware(dt) else dt
        return local.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return ""


def _nome_usuario_ajuste(aj) -> str:
    u = getattr(aj, "usuario", None)
    if u is None:
        return ""
    try:
        return (u.get_username() or "").strip()[:80]
    except Exception:
        return str(u)[:80]


def ultimos_ajustes_meta_por_produtos(pids: list[str]) -> dict[str, dict[str, dict[str, str]]]:
    """Último ajuste (qualquer origem) por depósito — data + quem."""
    out: dict[str, dict[str, dict[str, str]]] = {}
    uniq: list[str] = []
    seen: set[str] = set()
    for raw in pids:
        p = str(raw or "").strip()[:100]
        if not p or p in seen:
            continue
        seen.add(p)
        uniq.append(p)
        out[p] = {
            "centro": {"data": "", "quem": ""},
            "vila": {"data": "", "quem": ""},
        }
    if not uniq:
        return out

    qs = (
        AjusteRapidoEstoque.objects.filter(
            produto_externo_id__in=uniq,
            deposito__in=["centro", "vila"],
        )
        .select_related("usuario")
        .order_by("produto_externo_id", "deposito", "-criado_em")
        .only("produto_externo_id", "deposito", "criado_em", "usuario")
    )
    filled: set[tuple[str, str]] = set()
    for aj in qs.iterator(chunk_size=500):
        pid = str(aj.produto_externo_id or "").strip()
        dep = str(aj.deposito or "").strip().lower()
        if dep not in ("centro", "vila"):
            continue
        key = (pid, dep)
        if key in filled:
            continue
        filled.add(key)
        if pid not in out:
            out[pid] = {
                "centro": {"data": "", "quem": ""},
                "vila": {"data": "", "quem": ""},
            }
        out[pid][dep] = {
            "data": _fmt_dt_ajuste(getattr(aj, "criado_em", None)),
            "quem": _nome_usuario_ajuste(aj),
        }
    return out


def _enriquecer_saldos_e_meta(rows: list[dict]) -> list[dict]:
    if not rows:
        return rows
    from produtos.agro_fonte_config import agro_estoque_operacional_sem_mongo_erp
    from produtos.estoque_saldo_agro_util import mapa_saldos_operacionais_agro
    from produtos.views import obter_conexao_mongo

    p_ids = [str(r.get("id") or "").strip() for r in rows if r.get("id")]
    p_ids = [x for x in p_ids if x]
    if not p_ids:
        return rows

    client, db = (None, None)
    if not agro_estoque_operacional_sem_mongo_erp():
        try:
            client, db = obter_conexao_mongo()
        except Exception:
            client, db = None, None

    saldos = mapa_saldos_operacionais_agro(p_ids, db=db, client=client)
    metas = ultimos_ajustes_meta_por_produtos(p_ids)
    for r in rows:
        pid = str(r.get("id") or "").strip()
        s = saldos.get(pid) or {}
        sc = float(s.get("saldo_centro") or 0)
        sv = float(s.get("saldo_vila") or 0)
        r["saldo_centro"] = round(sc, 3)
        r["saldo_vila"] = round(sv, 3)
        r["saldo_total"] = round(sc + sv, 3)
        r["saldo_erp_centro"] = float(s.get("saldo_erp_centro") or 0)
        r["saldo_erp_vila"] = float(s.get("saldo_erp_vila") or 0)
        m = metas.get(pid) or {}
        r["ult_data_centro"] = (m.get("centro") or {}).get("data") or ""
        r["ult_quem_centro"] = (m.get("centro") or {}).get("quem") or ""
        r["ult_data_vila"] = (m.get("vila") or {}).get("data") or ""
        r["ult_quem_vila"] = (m.get("vila") or {}).get("quem") or ""
    return rows


def coletar_linhas_export_estoque(
    *,
    filtros: dict | None = None,
    inativos: bool = False,
    q: str = "",
) -> tuple[list[dict], bool]:
    """Linhas do catálogo com saldo + meta de última alteração, respeitando filtros do cadastro."""
    from produtos.agro_fonte_config import agro_catalogo_usa_postgres
    from produtos.cadastro_filtros_util import (
        filtros_cadastro_ativos,
        row_passa_filtros_cadastro,
    )

    filtros = dict(filtros or {})
    filtros["incluir_saldo"] = True
    rows: list[dict] = []
    truncado = False
    tem_filtros = filtros_cadastro_ativos(filtros)
    q = str(q or "").strip()

    if agro_catalogo_usa_postgres():
        from produtos import catalogo_agro

        if q:
            lim = min(EXPORT_MAX_ROWS, 2000)
            if tem_filtros:
                chunk = catalogo_agro.buscar_gestao(
                    q,
                    limit=lim,
                    status_q="todos" if inativos else "ativos",
                    filtros=filtros,
                )
            else:
                chunk = catalogo_agro.buscar(q, limit=lim, inativos=inativos)
            rows = list(chunk or [])
            rows = _enriquecer_saldos_e_meta(rows)
            if filtros.get("estoque_sinal"):
                rows = [r for r in rows if row_passa_filtros_cadastro(r, filtros)]
            truncado = len(rows) >= lim
            return rows[:EXPORT_MAX_ROWS], truncado

        pagina = 1
        por_pagina = 500
        while len(rows) < EXPORT_MAX_ROWS:
            chunk, has_more = catalogo_agro.listar_paginado(
                pagina=pagina,
                por_pagina=por_pagina,
                sort_key="nome",
                sort_direction=1,
                inativos=inativos,
                filtros=filtros if tem_filtros else None,
            )
            if not chunk:
                break
            enriched = _enriquecer_saldos_e_meta(list(chunk))
            if filtros.get("estoque_sinal"):
                # QS já filtra por sinal; revalida com saldo operacional fresco
                enriched = [r for r in enriched if row_passa_filtros_cadastro(r, filtros)]
            rows.extend(enriched)
            if not has_more:
                break
            pagina += 1
        truncado = len(rows) >= EXPORT_MAX_ROWS
        return rows[:EXPORT_MAX_ROWS], truncado

    # Mongo / espelho
    from produtos.views import (
        _CADASTRO_LISTA_MONGO_PROJ,
        _aplicar_produto_gestao_overlay_em_dict,
        _overlay_mapa_por_ids_chunked,
        _produto_mongo_para_cadastro_row,
        obter_conexao_mongo,
    )

    client, db = obter_conexao_mongo()
    if db is None:
        raise ValueError("Mongo indisponível — não foi possível exportar o estoque.")

    filtro_mongo: dict = {} if inativos else {"CadastroInativo": {"$ne": True}}
    cur = (
        db[client.col_p]
        .find(filtro_mongo, _CADASTRO_LISTA_MONGO_PROJ)
        .sort("Nome", 1)
        .limit(EXPORT_MAX_ROWS + 1)
    )
    chunk = list(cur)
    truncado = len(chunk) > EXPORT_MAX_ROWS
    chunk = chunk[:EXPORT_MAX_ROWS]
    raw_rows = [_produto_mongo_para_cadastro_row(p) for p in chunk]
    ovs = _overlay_mapa_por_ids_chunked([str(r.get("id") or "") for r in raw_rows])
    for r in raw_rows:
        _aplicar_produto_gestao_overlay_em_dict(r, ovs.get(str(r.get("id") or "")))
    f_sem_est = dict(filtros)
    f_sem_est["estoque_sinal"] = ""
    if tem_filtros:
        raw_rows = [r for r in raw_rows if row_passa_filtros_cadastro(r, f_sem_est)]
    raw_rows = _enriquecer_saldos_e_meta(raw_rows)
    if filtros.get("estoque_sinal"):
        raw_rows = [r for r in raw_rows if row_passa_filtros_cadastro(r, filtros)]
    if q:
        ql = q.casefold()
        raw_rows = [
            r
            for r in raw_rows
            if ql in str(r.get("nome") or "").casefold()
            or ql in str(r.get("codigo_nfe") or r.get("codigo") or "").casefold()
            or ql in str(r.get("codigo_barras") or "").casefold()
        ]
    return raw_rows[:EXPORT_MAX_ROWS], truncado


def linha_export_planilha_estoque(row: dict) -> dict[str, Any]:
    return {
        COL_ID: str(row.get("id") or ""),
        COL_CODIGO_GM: str(row.get("codigo_nfe") or row.get("codigo") or ""),
        COL_NOME: str(row.get("nome") or ""),
        COL_CODIGO_BARRAS: str(row.get("codigo_barras") or ""),
        COL_SALDO_CENTRO: float(row.get("saldo_centro") or 0),
        COL_AJUSTE_CENTRO: None,  # vazio — operador preenche se quiser delta
        COL_ULT_DATA_CENTRO: str(row.get("ult_data_centro") or ""),
        COL_ULT_QUEM_CENTRO: str(row.get("ult_quem_centro") or ""),
        COL_SALDO_VILA: float(row.get("saldo_vila") or 0),
        COL_AJUSTE_VILA: None,
        COL_ULT_DATA_VILA: str(row.get("ult_data_vila") or ""),
        COL_ULT_QUEM_VILA: str(row.get("ult_quem_vila") or ""),
    }


def montar_xlsx_estoque(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Estoque"
    hdr_fill = PatternFill("solid", fgColor="DBEAFE")
    hdr_font = Font(bold=True, color="1E3A8A")
    lock_fill = PatternFill("solid", fgColor="F1F5F9")
    edit_fill = PatternFill("solid", fgColor="FEF9C3")
    info_fill = PatternFill("solid", fgColor="F8FAFC")

    for col, (label, key) in enumerate(EXPORT_HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=label)
        c.font = hdr_font
        c.fill = hdr_fill
        c.protection = Protection(locked=True)

    for ri, src in enumerate(rows, start=2):
        line = linha_export_planilha_estoque(src)
        for col, (_, key) in enumerate(EXPORT_HEADERS, start=1):
            val = line.get(key)
            cell = ws.cell(row=ri, column=col)
            if key in (COL_AJUSTE_CENTRO, COL_AJUSTE_VILA):
                cell.value = None
            elif key in (COL_ID, COL_CODIGO_GM, COL_CODIGO_BARRAS) or key.startswith("ult_"):
                cell.value = str(val) if val is not None else ""
                cell.number_format = "@"
            elif key in (COL_SALDO_CENTRO, COL_SALDO_VILA):
                cell.value = val
                cell.number_format = "#,##0.###"
            else:
                cell.value = val

            bloqueada = key in EXPORT_COLS_BLOQUEADAS
            cell.protection = Protection(locked=bloqueada)
            if key in EXPORT_COLS_OCULTAS:
                cell.fill = lock_fill
            elif key in (COL_SALDO_CENTRO, COL_AJUSTE_CENTRO, COL_SALDO_VILA, COL_AJUSTE_VILA):
                cell.fill = edit_fill
            elif bloqueada:
                cell.fill = info_fill

    for col in range(1, len(EXPORT_HEADERS) + 1):
        letter = get_column_letter(col)
        key = EXPORT_HEADERS[col - 1][1]
        ws.column_dimensions[letter].width = 14
        if key in EXPORT_COLS_OCULTAS:
            ws.column_dimensions[letter].hidden = True
            ws.column_dimensions[letter].width = 2
        if key == COL_NOME:
            ws.column_dimensions[letter].width = 40
        if key.startswith("ult_"):
            ws.column_dimensions[letter].width = 16

    # Aba legenda
    ws2 = wb.create_sheet("Como usar")
    ws2["A1"] = "Como ajustar estoque nesta planilha"
    ws2["A1"].font = Font(bold=True, size=14)
    dicas = [
        "",
        "1) Filtre no Cadastro (saldo negativo/zerado/positivo, marca, categoria…) e baixe só o que precisa.",
        "2) Colunas amarelas são editáveis: Saldo Centro/Vila e Ajuste (+/−).",
        "3) Jeito mais fácil: deixe o Saldo como veio e digite só o Ajuste (+/−), ex.: +1 ou -2.",
        "4) Se Ajuste estiver preenchido, o sistema IGNORA o Saldo daquele depósito (atual + ajuste).",
        "5) Se quiser informar a quantidade final na prateleira, apague o Ajuste e altere só o Saldo.",
        "6) Sem Saldo nem Ajuste alterados naquele depósito = nada é gravado.",
        "7) Colunas «Últ. alt.» e «Quem» são só leitura — o sistema ignora se alterar.",
        "8) Não mexa na coluna ID (oculta). Apague linhas que não for ajustar.",
        "9) Ao subir o arquivo, confira a prévia antes de confirmar.",
    ]
    for i, t in enumerate(dicas, start=2):
        ws2.cell(row=i, column=1, value=t)
    ws2.column_dimensions["A"].width = 100

    ws.protection.sheet = True
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _patch_estoque_da_linha(raw: dict, colmap: dict[str, str | None]) -> dict[str, Any]:
    patch: dict[str, Any] = {}

    def dec_opt(key: str) -> None:
        hdr = colmap.get(key)
        if not hdr or hdr not in raw:
            return
        v = raw.get(hdr)
        if v is None or (isinstance(v, str) and not str(v).strip()):
            return
        d = _parse_decimal_br(v)
        if d is None:
            patch[f"__erro_{key}"] = f"Valor inválido em «{hdr}»."
        else:
            patch[key] = d

    dec_opt(COL_SALDO_CENTRO)
    dec_opt(COL_AJUSTE_CENTRO)
    dec_opt(COL_SALDO_VILA)
    dec_opt(COL_AJUSTE_VILA)
    return patch


def _resolver_alvo_deposito(
    atual: Decimal,
    saldo_abs: Decimal | None,
    ajuste: Decimal | None,
) -> tuple[Decimal | None, str | None]:
    """
    Resolve o saldo alvo do depósito.

    Prioridade (usabilidade loja):
    1) Se **Ajuste +/-** veio preenchido → usa ``atual + ajuste`` e **ignora** a coluna Saldo
       (a planilha exportada já traz o saldo atual; o operador só digita o +/-).
    2) Se só **Saldo** veio preenchido → usa o valor absoluto.
    3) Ambos vazios → sem alteração neste depósito.
    """
    if ajuste is not None:
        return _round3(atual + ajuste), None
    if saldo_abs is not None:
        return _round3(saldo_abs), None
    return None, None


def _estado_atual_estoque_mapa(
    pids: list[str],
    on_progress: ProgressCb = None,
) -> dict[str, dict]:
    from produtos.agro_fonte_config import agro_catalogo_usa_postgres
    from produtos.cadastro_planilha_util import _mapa_estado_atual_produtos

    if on_progress:
        on_progress(8, "Carregando produtos…")
    base = _mapa_estado_atual_produtos(pids, on_progress=on_progress)
    rows = []
    for pid, row in base.items():
        r = dict(row)
        r["id"] = pid
        rows.append(r)
    if on_progress:
        on_progress(20, "Calculando saldos…")
    enriched = _enriquecer_saldos_e_meta(rows)
    out = {str(r.get("id") or ""): r for r in enriched if r.get("id")}
    # produtos sem linha no mapa base mas com id pedido
    missing = [p for p in pids if p not in out]
    if missing and agro_catalogo_usa_postgres():
        # já tentou via mapa
        pass
    return out


def preview_importacao_estoque(path: Path, on_progress: ProgressCb = None) -> dict[str, Any]:
    headers, rows_raw = _ler_planilha(path)
    if on_progress:
        on_progress(0, f"total:{len(rows_raw)}")
        on_progress(2, f"Planilha com {len(rows_raw)} linha(s) — lendo…")
    colmap = _map_headers(headers)
    if not colmap.get(COL_ID):
        raise ValueError("Coluna «ID» não encontrada na planilha.")

    hdr_id = colmap[COL_ID]
    rows_slice = rows_raw[:IMPORT_MAX_ROWS]
    vistos: set[str] = set()
    alteracoes: list[dict] = []
    ignoradas: list[dict] = []
    erros: list[dict] = []
    pendentes: list[dict] = []

    for i, raw in enumerate(rows_slice, start=2):
        pid = _cel_str(raw.get(hdr_id or ""))[:64]
        if not pid:
            continue
        if not _id_produto_planilha_valido(pid):
            erros.append(
                {
                    "linha": i,
                    "id": _id_planilha_resumo(pid),
                    "erro": "ID inválido. Apague a linha ou restaure o ID da exportação.",
                }
            )
            continue
        if pid in vistos:
            erros.append({"linha": i, "id": pid, "erro": "ID duplicado na planilha."})
            continue
        vistos.add(pid)
        patch = _patch_estoque_da_linha(raw, colmap)
        err_fields = [v for k, v in patch.items() if k.startswith("__erro_")]
        if err_fields:
            erros.append({"linha": i, "id": pid, "erro": err_fields[0]})
            continue
        if not any(k in patch for k in IMPORT_EDIT_KEYS):
            ignoradas.append(
                {"linha": i, "id": pid, "motivo": "Sem Saldo nem Ajuste preenchido."}
            )
            continue
        pendentes.append({"linha": i, "id": pid, "patch": patch})

    if on_progress:
        on_progress(8, f"Conferindo {len(pendentes)} linha(s)…")

    mapa = _estado_atual_estoque_mapa([p["id"] for p in pendentes], on_progress=on_progress)
    total_pend = len(pendentes)
    step = max(1, total_pend // 40 or 1)

    for idx, item in enumerate(pendentes):
        pid = item["id"]
        patch = item["patch"]
        i = item["linha"]
        if on_progress and (idx == 0 or idx % step == 0 or idx == total_pend - 1):
            pct = 20 + int(75 * idx / max(1, total_pend))
            on_progress(pct, f"Analisando linha {i}… ({idx + 1}/{total_pend})")

        atual = mapa.get(pid)
        if not atual:
            erros.append({"linha": i, "id": pid, "erro": "Produto não encontrado no catálogo."})
            continue

        sc = Decimal(str(atual.get("saldo_centro") or 0))
        sv = Decimal(str(atual.get("saldo_vila") or 0))
        alvo_c, err_c = _resolver_alvo_deposito(
            sc, patch.get(COL_SALDO_CENTRO), patch.get(COL_AJUSTE_CENTRO)
        )
        alvo_v, err_v = _resolver_alvo_deposito(
            sv, patch.get(COL_SALDO_VILA), patch.get(COL_AJUSTE_VILA)
        )
        if err_c:
            erros.append({"linha": i, "id": pid, "erro": f"Centro: {err_c}"})
            continue
        if err_v:
            erros.append({"linha": i, "id": pid, "erro": f"Vila: {err_v}"})
            continue

        campos = []
        if alvo_c is not None and not _quase_igual(alvo_c, sc):
            campos.append(
                {
                    "campo": "centro",
                    "de": float(sc),
                    "para": float(alvo_c),
                    "delta": float(_round3(alvo_c - sc)),
                }
            )
        if alvo_v is not None and not _quase_igual(alvo_v, sv):
            campos.append(
                {
                    "campo": "vila",
                    "de": float(sv),
                    "para": float(alvo_v),
                    "delta": float(_round3(alvo_v - sv)),
                }
            )
        if not campos:
            ignoradas.append(
                {"linha": i, "id": pid, "motivo": "Saldo igual ao atual — nada a ajustar."}
            )
            continue

        alteracoes.append(
            {
                "linha": i,
                "id": pid,
                "nome": str(atual.get("nome") or ""),
                "campos": campos,
            }
        )

    if on_progress:
        on_progress(100, "Concluído")

    if len(rows_raw) > IMPORT_MAX_ROWS:
        erros.append(
            {
                "linha": None,
                "id": "",
                "erro": f"Planilha truncada: só as primeiras {IMPORT_MAX_ROWS} linhas foram analisadas.",
            }
        )

    return {
        "total_linhas": len(rows_raw),
        "alteracoes": alteracoes[:500],
        "n_alteracoes": len(alteracoes),
        "ignoradas": ignoradas[:80],
        "n_ignoradas": len(ignoradas),
        "erros": erros[:120],
        "n_erros": len(erros),
    }


def _empresa_padrao():
    from base.models import Empresa

    return Empresa.objects.filter(nome_fantasia="Agro Mais").first() or Empresa.objects.first()


def _gravar_ajuste_estoque(
    *,
    pid: str,
    nome: str,
    codigo: str,
    deposito: str,
    saldo_erp: Decimal,
    saldo_novo: Decimal,
    user,
) -> AjusteRapidoEstoque:
    return AjusteRapidoEstoque.objects.create(
        empresa=_empresa_padrao(),
        produto_externo_id=str(pid)[:100],
        codigo_interno=str(codigo or "")[:100],
        nome_produto=(nome or pid)[:255],
        deposito=deposito,
        saldo_erp_referencia=_round3(saldo_erp),
        saldo_informado=_round3(saldo_novo),
        origem=OrigemAjusteEstoque.PLANILHA,
        observacao="Cadastro — Excel estoque",
        usuario=user if user and getattr(user, "is_authenticated", False) else None,
    )


def aplicar_importacao_estoque(
    path: Path,
    user,
    *,
    nome_arquivo: str = "",
    on_progress: ProgressCb = None,
) -> dict[str, Any]:
    headers, rows_raw = _ler_planilha(path)
    colmap = _map_headers(headers)
    hdr_id = colmap.get(COL_ID)
    if not hdr_id:
        raise ValueError("Coluna «ID» não encontrada.")

    if on_progress:
        on_progress(2, f"total:{len(rows_raw)}")
        on_progress(4, f"Lendo {len(rows_raw)} linha(s)…")

    candidatos: list[dict] = []
    vistos: set[str] = set()
    for i, raw in enumerate(rows_raw[:IMPORT_MAX_ROWS], start=2):
        pid = _cel_str(raw.get(hdr_id or ""))[:64]
        if not pid or not _id_produto_planilha_valido(pid) or pid in vistos:
            continue
        vistos.add(pid)
        patch = _patch_estoque_da_linha(raw, colmap)
        if any(k.startswith("__erro_") for k in patch):
            continue
        if not any(k in patch for k in IMPORT_EDIT_KEYS):
            continue
        candidatos.append({"linha": i, "id": pid, "patch": patch})

    mapa = _estado_atual_estoque_mapa([c["id"] for c in candidatos], on_progress=on_progress)

    fila: list[dict] = []
    for item in candidatos:
        pid = item["id"]
        patch = item["patch"]
        i = item["linha"]
        atual = mapa.get(pid)
        if not atual:
            continue
        sc = Decimal(str(atual.get("saldo_centro") or 0))
        sv = Decimal(str(atual.get("saldo_vila") or 0))
        alvo_c, err_c = _resolver_alvo_deposito(
            sc, patch.get(COL_SALDO_CENTRO), patch.get(COL_AJUSTE_CENTRO)
        )
        alvo_v, err_v = _resolver_alvo_deposito(
            sv, patch.get(COL_SALDO_VILA), patch.get(COL_AJUSTE_VILA)
        )
        if err_c or err_v:
            continue
        deps = []
        if alvo_c is not None and not _quase_igual(alvo_c, sc):
            deps.append(
                {
                    "deposito": "centro",
                    "de": float(sc),
                    "para": float(alvo_c),
                    "erp": float(atual.get("saldo_erp_centro") or 0),
                }
            )
        if alvo_v is not None and not _quase_igual(alvo_v, sv):
            deps.append(
                {
                    "deposito": "vila",
                    "de": float(sv),
                    "para": float(alvo_v),
                    "erp": float(atual.get("saldo_erp_vila") or 0),
                }
            )
        if not deps:
            continue
        fila.append(
            {
                "linha": i,
                "id": pid,
                "nome": str(atual.get("nome") or ""),
                "codigo": str(atual.get("codigo_nfe") or atual.get("codigo") or ""),
                "deps": deps,
            }
        )

    if not fila:
        raise ValueError("Nenhum ajuste válido para gravar — confira a prévia.")

    if on_progress:
        on_progress(90, f"Gravando {len(fila)} produto(s)…")

    ok = 0
    falhas: list[dict] = []
    backups: list[dict] = []
    n_campos_total = 0
    total_fila = len(fila)

    for idx, item in enumerate(fila):
        if on_progress and (idx == 0 or idx == total_fila - 1 or idx % max(1, total_fila // 20) == 0):
            pct = 90 + int(8 * idx / max(1, total_fila))
            on_progress(pct, f"Gravando produto {idx + 1}/{total_fila}…")
        pid = item["id"]
        snap_deps = []
        try:
            with transaction.atomic():
                for d in item["deps"]:
                    aj = _gravar_ajuste_estoque(
                        pid=pid,
                        nome=item["nome"],
                        codigo=item["codigo"],
                        deposito=d["deposito"],
                        saldo_erp=Decimal(str(d["erp"])),
                        saldo_novo=Decimal(str(d["para"])),
                        user=user,
                    )
                    snap_deps.append(
                        {
                            "deposito": d["deposito"],
                            "de": d["de"],
                            "para": d["para"],
                            "erp": d["erp"],
                            "ajuste_id": aj.pk,
                        }
                    )
            backups.append(
                {
                    "id": pid,
                    "nome": item["nome"],
                    "codigo": item["codigo"],
                    "deps": snap_deps,
                    "campos_alterados": [d["deposito"] for d in snap_deps],
                }
            )
            n_campos_total += len(snap_deps)
            ok += 1
        except Exception as exc:
            falhas.append({"linha": item["linha"], "id": pid, "erro": str(exc) or "Falha."})

    historico_id = None
    if backups:
        create_kw = dict(
            usuario=user if user and getattr(user, "is_authenticated", False) else None,
            nome_arquivo=str(nome_arquivo or "")[:255],
            n_produtos=len(backups),
            n_campos=n_campos_total,
            backup={"tipo": TIPO_HISTORICO, "items": backups},
        )
        # Campo tipo se existir (migration)
        if hasattr(CadastroPlanilhaImportHistoricoAgro, "tipo"):
            create_kw["tipo"] = TIPO_HISTORICO
        hist = CadastroPlanilhaImportHistoricoAgro.objects.create(**create_kw)
        historico_id = hist.pk

    if ok:
        try:
            from produtos.views import _invalidar_caches_apos_ajuste_pin

            _invalidar_caches_apos_ajuste_pin()
        except Exception:
            pass
        # Invalida snapshot de saldos / catálogo leve
        try:
            cache.delete("agro_pdv_saldos_snapshot_v1")
        except Exception:
            pass

    if on_progress:
        on_progress(100, "Concluído")

    return {
        "gravados": ok,
        "falhas": falhas[:80],
        "n_falhas": len(falhas),
        "historico_id": historico_id,
        "n_alteracoes": len(fila),
    }


def listar_historico_import_estoque(*, limite: int = HISTORICO_LISTA_LIMITE) -> list[dict]:
    out: list[dict] = []
    qs = CadastroPlanilhaImportHistoricoAgro.objects.select_related(
        "usuario", "revertido_por"
    ).order_by("-criado_em")
    if hasattr(CadastroPlanilhaImportHistoricoAgro, "tipo"):
        qs = qs.filter(tipo=TIPO_HISTORICO)
    else:
        # Fallback pré-migration: só backups marcados
        qs = qs.all()
    for h in qs[: limite * 3]:
        backup = h.backup or {}
        if backup.get("tipo") != TIPO_HISTORICO and getattr(h, "tipo", "") != TIPO_HISTORICO:
            continue
        items = backup.get("items") or []
        out.append(
            {
                "id": h.pk,
                "criado_em": h.criado_em.isoformat() if h.criado_em else "",
                "usuario": (h.usuario.get_username() if h.usuario else "") or "",
                "nome_arquivo": h.nome_arquivo or "",
                "n_produtos": h.n_produtos,
                "n_campos": h.n_campos,
                "status": h.status,
                "pode_reverter": h.status == CadastroPlanilhaImportHistoricoAgro.Status.APLICADO,
                "revertido_em": h.revertido_em.isoformat() if h.revertido_em else "",
                "revertido_por": (h.revertido_por.get_username() if h.revertido_por else "")
                or "",
                "resumo": [
                    {
                        "id": it.get("id"),
                        "nome": it.get("nome") or "",
                        "campos": it.get("campos_alterados") or [],
                        "detalhes": [
                            {
                                "campo": d.get("deposito"),
                                "de": d.get("de"),
                                "para": d.get("para"),
                            }
                            for d in (it.get("deps") or [])[:4]
                        ],
                    }
                    for it in items[:12]
                ],
            }
        )
        if len(out) >= limite:
            break
    return out


def reverter_importacao_estoque(historico_id: int, user) -> dict[str, Any]:
    hist = CadastroPlanilhaImportHistoricoAgro.objects.filter(pk=historico_id).first()
    if not hist:
        raise ValueError("Histórico não encontrado.")
    backup = hist.backup or {}
    if backup.get("tipo") != TIPO_HISTORICO and getattr(hist, "tipo", "") != TIPO_HISTORICO:
        raise ValueError("Este histórico não é de Excel estoque.")
    if hist.status != CadastroPlanilhaImportHistoricoAgro.Status.APLICADO:
        raise ValueError("Esta importação já foi desfeita.")

    items = backup.get("items") or []
    if not items:
        raise ValueError("Backup vazio — não é possível desfazer.")

    from produtos.agro_fonte_config import agro_estoque_operacional_sem_mongo_erp
    from produtos.estoque_saldo_agro_util import mapa_saldos_operacionais_agro
    from produtos.views import obter_conexao_mongo

    pids = [str(it.get("id") or "").strip() for it in items if it.get("id")]
    client, db = (None, None)
    if not agro_estoque_operacional_sem_mongo_erp():
        try:
            client, db = obter_conexao_mongo()
        except Exception:
            client, db = None, None
    saldos = mapa_saldos_operacionais_agro(pids, db=db, client=client)

    with transaction.atomic():
        for item in items:
            pid = str(item.get("id") or "").strip()
            if not pid:
                continue
            sinfo = saldos.get(pid) or {}
            for d in item.get("deps") or []:
                dep = str(d.get("deposito") or "").strip().lower()
                if dep not in ("centro", "vila"):
                    continue
                erp_key = "saldo_erp_centro" if dep == "centro" else "saldo_erp_vila"
                erp = Decimal(str(sinfo.get(erp_key) or d.get("erp") or 0))
                _gravar_ajuste_estoque(
                    pid=pid,
                    nome=str(item.get("nome") or pid),
                    codigo=str(item.get("codigo") or ""),
                    deposito=dep,
                    saldo_erp=erp,
                    saldo_novo=Decimal(str(d.get("de") or 0)),
                    user=user,
                )
        hist.status = CadastroPlanilhaImportHistoricoAgro.Status.REVERTIDO
        hist.revertido_em = timezone.now()
        hist.revertido_por = user if user and getattr(user, "is_authenticated", False) else None
        hist.save(update_fields=["status", "revertido_em", "revertido_por"])

    try:
        from produtos.views import _invalidar_caches_apos_ajuste_pin

        _invalidar_caches_apos_ajuste_pin()
    except Exception:
        pass
    try:
        cache.delete("agro_pdv_saldos_snapshot_v1")
    except Exception:
        pass

    return {
        "historico_id": hist.pk,
        "revertidos": len(items),
        "status": hist.status,
    }

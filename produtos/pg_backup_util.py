"""Export/import portável Postgres Agro — ZIP (JSONL) + Excel resumo (FL-048)."""
from __future__ import annotations

import hashlib
import io
import json
import zipfile
from datetime import datetime, timezone as dt_timezone
from typing import Any, BinaryIO, Iterator

from django.apps import apps
from django.conf import settings
from django.core import serializers
from django.core.serializers.base import DeserializationError
from django.db import transaction
from django.db.models import Model
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from config.app_build_util import read_app_version
from produtos.pg_backup_registry import (
    PG_BACKUP_ALL_SLUGS,
    PG_BACKUP_CATEGORIES,
    PG_BACKUP_CATEGORY_BY_SLUG,
    PG_BACKUP_FORMAT,
    PgBackupCategory,
)

_EXCEL_PREVIEW_MAX_ROWS = 150
_EXCEL_FULL_MAX_ROWS = 2500
_JSONL_CHUNK = 500


def _utc_now_iso() -> str:
    return timezone.now().astimezone(dt_timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _get_model(label: str) -> type[Model]:
    app_label, model_name = label.split(".", 1)
    return apps.get_model(app_label, model_name)


def _model_row_count(label: str) -> int:
    try:
        return _get_model(label).objects.count()
    except Exception:
        return -1


def _category_stats(cat: PgBackupCategory) -> dict[str, Any]:
    models_stats = []
    total = 0
    for label in cat.models:
        n = _model_row_count(label)
        if n > 0:
            total += n
        models_stats.append({"label": label, "count": n})
    return {
        "slug": cat.slug,
        "label": cat.label,
        "warning": cat.warning,
        "total_rows": total,
        "models": models_stats,
    }


def listar_categorias_stats() -> list[dict[str, Any]]:
    return [_category_stats(c) for c in PG_BACKUP_CATEGORIES]


def _normalize_slugs(slugs: list[str] | None) -> list[str]:
    if not slugs:
        return list(PG_BACKUP_ALL_SLUGS)
    out = []
    seen = set()
    for s in slugs:
        s = (s or "").strip()
        if not s or s in seen:
            continue
        if s not in PG_BACKUP_CATEGORY_BY_SLUG:
            continue
        seen.add(s)
        out.append(s)
    order = {slug: i for i, slug in enumerate(PG_BACKUP_ALL_SLUGS)}
    out.sort(key=lambda x: order.get(x, 999))
    return out


def _serialize_queryset_jsonl(qs) -> Iterator[bytes]:
    for obj in qs.iterator(chunk_size=_JSONL_CHUNK):
        line = serializers.serialize("json", [obj])
        inner = json.loads(line)
        if inner:
            yield (json.dumps(inner[0], ensure_ascii=False) + "\n").encode("utf-8")


def _flat_row_from_instance(obj: Model, max_fields: int = 24) -> dict[str, Any]:
    row: dict[str, Any] = {"_modelo": obj._meta.label, "_pk": obj.pk}
    for i, field in enumerate(obj._meta.fields):
        if i >= max_fields:
            row["_truncado"] = True
            break
        val = getattr(obj, field.name)
        if val is None:
            row[field.name] = ""
        elif hasattr(val, "isoformat"):
            row[field.name] = val.isoformat()
        else:
            s = str(val)
            row[field.name] = s[:500] if len(s) > 500 else s
    return row


def _build_excel_resumo(
    *,
    manifest: dict[str, Any],
    categories: list[PgBackupCategory],
    export_stats: dict[str, Any],
) -> bytes:
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Manifesto"
    ws0.append(["Campo", "Valor"])
    for key in (
        "format",
        "version_app",
        "gerado_em",
        "gerado_por",
        "ambiente",
        "categorias",
        "total_registros",
    ):
        ws0.append([key, manifest.get(key, "")])
    for cell in ws0[1]:
        cell.font = Font(bold=True)

    ws1 = wb.create_sheet("Contagens")
    ws1.append(["Categoria", "Modelo", "Registros exportados"])
    for cell in ws1[1]:
        cell.font = Font(bold=True)
    for cat in categories:
        for m in export_stats.get(cat.slug, {}).get("models", []):
            ws1.append([cat.label, m["label"], m["count"]])

    for cat in categories:
        safe = cat.slug[:28]
        ws = wb.create_sheet(safe)
        headers_written = False
        row_num = 0
        for label in cat.models:
            model = _get_model(label)
            qs = model.objects.all().order_by("pk")
            limit = _EXCEL_FULL_MAX_ROWS if model.objects.count() <= _EXCEL_FULL_MAX_ROWS else _EXCEL_PREVIEW_MAX_ROWS
            for obj in qs.iterator(chunk_size=200):
                if row_num >= limit:
                    ws.append(["…", f"(mais linhas só no JSONL — modelo {label})"])
                    break
                flat = _flat_row_from_instance(obj)
                if not headers_written:
                    ws.append(list(flat.keys()))
                    headers_written = True
                    for cell in ws[1]:
                        cell.font = Font(bold=True)
                ws.append([flat.get(k, "") for k in flat.keys()])
                row_num += 1
        if not headers_written:
            ws.append(["(vazio)", ""])

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 8
            col_letter = get_column_letter(col[0].column)
            for cell in col[:80]:
                if cell.value is not None:
                    max_len = max(max_len, min(48, len(str(cell.value))))
            sheet.column_dimensions[col_letter].width = max_len + 2

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_backup_zip(
    slugs: list[str] | None,
    *,
    username: str = "",
) -> tuple[bytes, str, dict[str, Any]]:
    selected_slugs = _normalize_slugs(slugs)
    if not selected_slugs:
        raise ValueError("Nenhuma categoria válida selecionada.")

    categories = [PG_BACKUP_CATEGORY_BY_SLUG[s] for s in selected_slugs]
    export_stats: dict[str, Any] = {}
    total_records = 0
    sha = hashlib.sha256()

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for cat in categories:
            cat_stats = {"models": [], "bytes": 0}
            cat_buf = io.BytesIO()
            for label in cat.models:
                model = _get_model(label)
                cnt = model.objects.count()
                cat_stats["models"].append({"label": label, "count": cnt})
                total_records += cnt
                for chunk in _serialize_queryset_jsonl(model.objects.all().order_by("pk")):
                    cat_buf.write(chunk)
                    sha.update(chunk)
            data = cat_buf.getvalue()
            cat_stats["bytes"] = len(data)
            export_stats[cat.slug] = cat_stats
            zf.writestr(f"data/{cat.slug}.jsonl", data)

        manifest = {
            "format": PG_BACKUP_FORMAT,
            "version_app": read_app_version(),
            "gerado_em": _utc_now_iso(),
            "gerado_por": username or "",
            "ambiente": getattr(settings, "AGRO_DEPLOY_AMBIENTE", "") or "desconhecido",
            "categorias": selected_slugs,
            "total_registros": total_records,
            "sha256_payload": "",  # filled below
        }
        manifest["sha256_payload"] = sha.hexdigest()
        manifest_bytes = json.dumps(manifest, ensure_ascii=False, indent=2).encode("utf-8")
        zf.writestr("manifest.json", manifest_bytes)
        zf.writestr(
            "resumo.xlsx",
            _build_excel_resumo(manifest=manifest, categories=categories, export_stats=export_stats),
        )

    stamp = timezone.now().strftime("%Y%m%d-%H%M%S")
    scope = "completo" if len(selected_slugs) == len(PG_BACKUP_ALL_SLUGS) else "parcial"
    filename = f"sistvale-pg-backup-{scope}-{stamp}.zip"
    return buf.getvalue(), filename, manifest


def _delete_category_models(cat: PgBackupCategory) -> dict[str, int]:
    deleted: dict[str, int] = {}
    for label in reversed(cat.models):
        model = _get_model(label)
        n, _ = model.objects.all().delete()
        deleted[label] = n
    return deleted


def _load_jsonl_for_category(cat: PgBackupCategory, raw: bytes) -> dict[str, Any]:
    stream = io.BytesIO(raw)
    text_stream = io.TextIOWrapper(stream, encoding="utf-8")
    loaded = 0
    errors: list[str] = []
    model_order = {label: i for i, label in enumerate(cat.models)}
    pending: list[tuple[int, str, Any]] = []

    for line_no, line in enumerate(text_stream, start=1):
        line = line.strip()
        if not line:
            continue
        try:
            wrapper = json.loads(line)
            model_label = wrapper.get("model", "")
            pending.append((model_order.get(model_label, 999), line, model_label))
        except json.JSONDecodeError as exc:
            errors.append(f"Linha {line_no}: JSON inválido ({exc})")

    pending.sort(key=lambda x: x[0])

    for _, line, model_label in pending:
        try:
            for obj in serializers.deserialize("jsonl", line + "\n"):
                obj.save()
                loaded += 1
        except (DeserializationError, Exception) as exc:
            errors.append(f"{model_label or '?'}: {exc}")

    return {"loaded": loaded, "errors": errors}


def restore_backup_zip(
    uploaded: BinaryIO,
    slugs_filter: list[str] | None,
    *,
    username: str = "",
) -> dict[str, Any]:
    selected = set(_normalize_slugs(slugs_filter)) if slugs_filter else None
    result: dict[str, Any] = {
        "ok": True,
        "categorias": [],
        "erros": [],
        "usuario": username,
    }

    with zipfile.ZipFile(uploaded, "r") as zf:
        try:
            manifest = json.loads(zf.read("manifest.json").decode("utf-8"))
        except Exception as exc:
            return {"ok": False, "erros": [f"manifest.json inválido: {exc}"]}

        if manifest.get("format") != PG_BACKUP_FORMAT:
            result["erros"].append(
                f"Formato não suportado: {manifest.get('format')!r} (esperado {PG_BACKUP_FORMAT})"
            )
            result["ok"] = False

        zip_cats = manifest.get("categorias") or []
        to_restore = []
        for slug in zip_cats:
            if slug not in PG_BACKUP_CATEGORY_BY_SLUG:
                result["erros"].append(f"Categoria desconhecida no ZIP: {slug}")
                continue
            if selected is not None and slug not in selected:
                continue
            to_restore.append(PG_BACKUP_CATEGORY_BY_SLUG[slug])

        order = {slug: i for i, slug in enumerate(PG_BACKUP_ALL_SLUGS)}
        to_restore.sort(key=lambda c: order.get(c.slug, 999))

        if not to_restore:
            result["ok"] = False
            result["erros"].append("Nenhuma categoria para restaurar.")
            return result

        for cat in to_restore:
            path = f"data/{cat.slug}.jsonl"
            if path not in zf.namelist():
                result["erros"].append(f"Arquivo ausente no ZIP: {path}")
                result["ok"] = False
                continue

            cat_result = {"slug": cat.slug, "label": cat.label}
            try:
                with transaction.atomic():
                    cat_result["deleted"] = _delete_category_models(cat)
                    raw = zf.read(path)
                    load_info = _load_jsonl_for_category(cat, raw)
                    cat_result.update(load_info)
                    if load_info["errors"]:
                        raise RuntimeError("; ".join(load_info["errors"][:3]))
            except Exception as exc:
                result["ok"] = False
                cat_result["fatal"] = str(exc)
                result["erros"].append(f"{cat.label}: {exc}")
            result["categorias"].append(cat_result)

    return result

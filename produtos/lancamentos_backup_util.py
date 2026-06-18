"""Backup completo de lançamentos (pré-corte ERP) — Excel com todos os títulos."""
from __future__ import annotations

from io import BytesIO
from typing import Any

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from produtos.mongo_financeiro_util import (
    AGRO_CONGELADO_EM,
    AGRO_FONTE_VERDADE,
    COL_DTO_LANCAMENTO,
    lancamento_para_api,
)

_BACKUP_CAP = 120_000


def _linha_backup(doc: dict, despesa: bool) -> dict[str, Any]:
    row = lancamento_para_api(doc, despesa)
    row["tipo_label"] = "A pagar" if despesa else "A receber"
    row["fonte_agro"] = "Sim" if doc.get(AGRO_FONTE_VERDADE) else "Não"
    ce = doc.get(AGRO_CONGELADO_EM)
    row["congelado_em"] = (str(ce)[:19] if ce else "")
    row["criado_por"] = str(doc.get("CriadoPor") or "")[:200]
    row["modificado_por"] = str(doc.get("ModificadoPor") or "")[:200]
    row["id_erp"] = str(doc.get("Id") or doc.get("ID") or "")[:80]
    return row


def _filtro_mongo_backup(despesa: bool) -> dict[str, Any]:
    """Todos os documentos do tipo — sem dedup da lista (backup = espelho inteiro)."""
    if despesa:
        return {"Despesa": True}
    return {
        "$or": [
            {"Despesa": False},
            {"Despesa": {"$exists": False}},
            {"Despesa": None},
        ]
    }


def _coletar_mongo_backup_bruto(db, despesa: bool, cap: int) -> tuple[list[dict], int]:
    col = db[COL_DTO_LANCAMENTO]
    filtro = _filtro_mongo_backup(despesa)
    total = int(col.count_documents(filtro))
    cur = col.find(filtro).sort([("DataVencimento", 1), ("_id", 1)]).limit(max(1, cap))
    return [_linha_backup(d, despesa) for d in cur], total


def _coletar_fiado_backup_postgres() -> tuple[list[dict], int]:
    """Fiado PDV (Postgres) — aba separada; não altera o módulo fiado."""
    from produtos.fiado_gestao_util import titulo_para_dict
    from produtos.models import FiadoTituloAgro

    qs = FiadoTituloAgro.objects.select_related("cliente_agro", "venda_agro").order_by("pk")
    total = qs.count()
    return [titulo_para_dict(t) for t in qs[:_BACKUP_CAP]], total


def lancamentos_coletar_backup_completo(db) -> dict[str, Any]:
    """Pagar + receber (Mongo bruto) + fiado PDV (Postgres), para arquivo único."""
    if db is None:
        return {"ok": False, "erro": "Mongo indisponível"}
    cap = _BACKUP_CAP
    pagar, total_p = _coletar_mongo_backup_bruto(db, True, cap)
    receber, total_r = _coletar_mongo_backup_bruto(db, False, cap)
    fiado, total_fiado = _coletar_fiado_backup_postgres()

    try:
        col = db[COL_DTO_LANCAMENTO]
        mongo_bruto_pagar = int(col.count_documents({"Despesa": True}))
        mongo_bruto_receber = int(col.count_documents(_filtro_mongo_backup(False)))
    except Exception:
        mongo_bruto_pagar = None
        mongo_bruto_receber = None

    return {
        "ok": True,
        "pagar": pagar,
        "receber": receber,
        "fiado": fiado,
        "total_pagar": total_p,
        "total_receber": total_r,
        "total_fiado": total_fiado,
        "mongo_bruto_pagar": mongo_bruto_pagar,
        "mongo_bruto_receber": mongo_bruto_receber,
        "truncado": total_p > len(pagar) or total_r > len(receber) or total_fiado > len(fiado),
        "limite": cap,
    }


def _escrever_aba_lancamentos(ws, linhas: list[dict], *, despesa: bool, fiado_qtd: int = 0) -> None:
    head_font = Font(bold=True)
    label_mov = "Pago" if despesa else "Recebido"
    label_saldo = "A pagar" if despesa else "A receber"
    if not linhas:
        tit = "A pagar" if despesa else "A receber"
        ws["A1"] = f"Nenhum título «{tit}» no espelho financeiro (Mongo/ERP)."
        ws["A1"].font = head_font
        partes = [
            "Backup inclui todos os documentos deste tipo — não usa filtro da tela.",
        ]
        if not despesa and fiado_qtd > 0:
            partes.append(
                f"A loja usa Fiado no PDV ({fiado_qtd} título(s) na aba «Fiado PDV»). "
                "Isso não entra em Contas a receber do financeiro legado."
            )
        else:
            partes.append(
                "Se a loja só cobra fiado no caixa, esta aba pode ficar vazia mesmo assim."
            )
        ws["A2"] = " ".join(partes)
        ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
        return
    headers = (
        "ID SisVale",
        "ID ERP",
        "Vencimento",
        "Competência",
        "Data pagamento",
        "Cliente / favorecido",
        "Descrição",
        "Documento",
        "Parcela",
        "Plano conta",
        "Grupo",
        "Forma pagamento",
        "Banco",
        "Centro custo",
        "Empresa",
        "Valor bruto",
        label_mov,
        label_saldo,
        "Situação",
        "Observações",
        "Carimbo SisVale",
        "Congelado em",
        "Criado por",
        "Alterado por",
    )
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = head_font
        c.alignment = Alignment(wrap_text=True, vertical="top")
    for ri, row in enumerate(linhas, start=2):
        vals = (
            row.get("id") or "",
            row.get("id_erp") or "",
            (row.get("data_vencimento") or "")[:10],
            (row.get("data_competencia") or "")[:10],
            (row.get("data_pagamento") or "")[:10],
            row.get("cliente") or "",
            row.get("descricao") or "",
            row.get("numero_documento") or "",
            row.get("parcela") or 0,
            row.get("plano_conta") or "",
            row.get("grupo") or "",
            row.get("forma_pagamento") or "",
            row.get("banco") or "",
            row.get("centro_custo") or "",
            row.get("empresa") or "",
            row.get("valor_bruto"),
            row.get("valor_movimentado"),
            row.get("restante"),
            "Quitado" if row.get("pago") else "Aberto",
            row.get("observacoes") or "",
            row.get("fonte_agro") or "",
            row.get("congelado_em") or "",
            row.get("criado_por") or "",
            row.get("modificado_por") or "",
        )
        for col, v in enumerate(vals, start=1):
            ws.cell(row=ri, column=col, value=v)


def _escrever_aba_fiado(ws, linhas: list[dict]) -> None:
    head_font = Font(bold=True)
    ws["A1"] = (
        "Fiado / crédito loja (PDV) — Postgres SisVale. "
        "Aba separada; não é Contas a receber do financeiro legado."
    )
    ws["A1"].font = head_font
    ws.merge_cells("A1:L1")
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    if not linhas:
        ws["A3"] = "Nenhum título de fiado cadastrado."
        return
    headers = (
        "ID",
        "Cliente",
        "Código cliente",
        "Documento",
        "Parcela",
        "Vencimento",
        "Valor",
        "Pago",
        "Saldo",
        "Situação",
        "Origem",
        "Venda Agro ID",
        "Descrição",
    )
    start = 3
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=start, column=col, value=h)
        c.font = head_font
    for ri, row in enumerate(linhas, start=start + 1):
        vals = (
            row.get("id"),
            row.get("cliente_nome") or "",
            row.get("cliente_codigo") or "",
            row.get("numero_documento") or "",
            f"{row.get('parcela_num') or 0}/{row.get('parcela_total') or 0}",
            row.get("vencimento_texto") or (row.get("vencimento") or "")[:10],
            row.get("valor_bruto"),
            row.get("valor_pago"),
            row.get("saldo_aberto"),
            row.get("situacao_label") or row.get("situacao") or "",
            row.get("origem") or "",
            row.get("venda_agro_id") or "",
            row.get("descricao") or "",
        )
        for col, v in enumerate(vals, start=1):
            ws.cell(row=ri, column=col, value=v)


def montar_xlsx_backup_completo(
    dados: dict[str, Any],
    *,
    gerado_por: str = "",
) -> bytes:
    fiado_rows = dados.get("fiado") or []
    wb = Workbook()
    ws_r = wb.active
    ws_r.title = "A receber"
    _escrever_aba_lancamentos(ws_r, dados.get("receber") or [], despesa=False, fiado_qtd=len(fiado_rows))
    ws_p = wb.create_sheet("A pagar")
    _escrever_aba_lancamentos(ws_p, dados.get("pagar") or [], despesa=True)
    ws_f = wb.create_sheet("Fiado PDV")
    _escrever_aba_fiado(ws_f, fiado_rows)
    ws_i = wb.create_sheet("Resumo")
    agora = timezone.localtime()
    bold = Font(bold=True)
    ws_i["A1"] = "Backup completo — Lançamentos + Fiado (referência)"
    ws_i["A1"].font = bold
    ws_i["A2"] = "Gerado em"
    ws_i["B2"] = agora.strftime("%d/%m/%Y %H:%M")
    ws_i["A3"] = "Por"
    ws_i["B3"] = (gerado_por or "—")[:120]
    ws_i["A5"] = "A pagar (Mongo, todos os docs)"
    ws_i["B5"] = dados.get("total_pagar", 0)
    ws_i["A6"] = "A receber (Mongo, todos os docs)"
    ws_i["B6"] = dados.get("total_receber", 0)
    ws_i["A7"] = "Fiado PDV (Postgres — módulo separado)"
    ws_i["B7"] = dados.get("total_fiado", 0)
    ws_i["A8"] = "Linhas exportadas (pagar)"
    ws_i["B8"] = len(dados.get("pagar") or [])
    ws_i["A9"] = "Linhas exportadas (receber)"
    ws_i["B9"] = len(dados.get("receber") or [])
    ws_i["A10"] = "Linhas exportadas (fiado)"
    ws_i["B10"] = len(fiado_rows)
    if dados.get("truncado"):
        ws_i["A12"] = "Atenção"
        ws_i["B12"] = (
            f"Alguma aba atingiu o limite de {dados.get('limite')} linhas. "
            "Avise o suporte se faltar dado."
        )
    ws_i["A14"] = "Uso"
    ws_i["B14"] = (
        "Guarde no PC antes do checkpoint/corte ERP. "
        "Fiado continua no PDV como hoje — esta aba é só cópia de segurança."
    )
    ws_i["B14"].alignment = Alignment(wrap_text=True, vertical="top")
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def nome_arquivo_backup_completo() -> str:
    agora = timezone.localtime()
    return f"SisVale_Lancamentos_BACKUP_{agora.strftime('%Y-%m-%d_%H%M')}.xlsx"

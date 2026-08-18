"""Agregações e Excel da Central de Relatórios (fonte VendaAgro / ItemVendaAgro)."""
from __future__ import annotations

import logging
from collections import defaultdict
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from io import BytesIO
from typing import Any

from django.db.models import Max, Q, Sum
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def _aware_bounds(desde: datetime, ate: datetime) -> tuple[datetime, datetime]:
    if timezone.is_naive(desde):
        desde = timezone.make_aware(desde, timezone.get_current_timezone())
    if timezone.is_naive(ate):
        ate = timezone.make_aware(ate, timezone.get_current_timezone())
    return desde, ate


def parse_periodo_request(
    request,
    *,
    padrao: str = "mes_atual",
) -> dict[str, Any]:
    """
    Lê GET: periodo (hoje|7d|30d|mes_atual|custom) + de/ate (YYYY-MM-DD).
    Se De/Até vierem e não baterem com o atalho do período, usa personalizado.
    """
    hoje = timezone.localdate()
    periodo = (request.GET.get("periodo") or padrao).strip().lower()
    de_s = (request.GET.get("de") or "").strip()
    ate_s = (request.GET.get("ate") or "").strip()

    def _parse_d(s: str) -> date | None:
        try:
            return date.fromisoformat(s[:10])
        except (TypeError, ValueError):
            return None

    d_de = _parse_d(de_s)
    d_ate = _parse_d(ate_s)

    if periodo == "hoje":
        d0, d1 = hoje, hoje
    elif periodo == "7d":
        d0, d1 = hoje - timedelta(days=6), hoje
    elif periodo == "30d":
        d0, d1 = hoje - timedelta(days=29), hoje
    elif periodo == "custom":
        d0 = d_de or (hoje - timedelta(days=29))
        d1 = d_ate or hoje
        if d0 > d1:
            d0, d1 = d1, d0
    else:
        periodo = "mes_atual"
        d0 = hoje.replace(day=1)
        d1 = hoje

    # Usuário mudou as datas sem trocar o select → respeita De/Até (vira personalizado)
    if d_de and d_ate and periodo != "custom" and (d_de != d0 or d_ate != d1):
        periodo = "custom"
        d0, d1 = d_de, d_ate
        if d0 > d1:
            d0, d1 = d1, d0

    desde = datetime.combine(d0, time.min)
    ate = datetime.combine(d1, time(23, 59, 59))
    desde, ate = _aware_bounds(desde, ate)
    return {
        "periodo": periodo,
        "de": d0.isoformat(),
        "ate": d1.isoformat(),
        "desde": desde,
        "ate_dt": ate,
        "label": f"{d0.strftime('%d/%m/%Y')} — {d1.strftime('%d/%m/%Y')}",
    }


def parse_periodo_b_request(request) -> dict[str, Any]:
    """Segundo período para comparativo (de_b / ate_b ou atalho mes_passado)."""
    hoje = timezone.localdate()
    modo = (request.GET.get("periodo_b") or "mes_passado").strip().lower()
    de_s = (request.GET.get("de_b") or "").strip()
    ate_s = (request.GET.get("ate_b") or "").strip()

    def _parse_d(s: str) -> date | None:
        try:
            return date.fromisoformat(s[:10])
        except (TypeError, ValueError):
            return None

    d_de = _parse_d(de_s)
    d_ate = _parse_d(ate_s)

    if modo == "custom":
        d0 = d_de
        d1 = d_ate
        if d0 is None or d1 is None:
            primeiro = hoje.replace(day=1)
            d1 = primeiro - timedelta(days=1)
            d0 = d1.replace(day=1)
    else:
        primeiro = hoje.replace(day=1)
        d1 = primeiro - timedelta(days=1)
        d0 = d1.replace(day=1)
        modo = "mes_passado"

    if d_de and d_ate and modo != "custom" and (d_de != d0 or d_ate != d1):
        modo = "custom"
        d0, d1 = d_de, d_ate

    if d0 > d1:
        d0, d1 = d1, d0
    desde = datetime.combine(d0, time.min)
    ate = datetime.combine(d1, time(23, 59, 59))
    desde, ate = _aware_bounds(desde, ate)
    return {
        "periodo_b": modo,
        "de_b": d0.isoformat(),
        "ate_b": d1.isoformat(),
        "desde": desde,
        "ate_dt": ate,
        "label": f"{d0.strftime('%d/%m/%Y')} — {d1.strftime('%d/%m/%Y')}",
    }


def _qs_itens(desde: datetime, ate: datetime, deposito: str | None = None):
    from produtos.models import ItemVendaAgro

    qs = ItemVendaAgro.objects.filter(
        venda__devolvida_em__isnull=True,
        venda__criado_em__gte=desde,
        venda__criado_em__lte=ate,
    ).exclude(produto_id_externo="")
    if deposito == "vila":
        qs = qs.filter(venda__deposito__iexact="vila")
    elif deposito == "centro":
        qs = qs.filter(
            Q(venda__deposito__iexact="centro")
            | Q(venda__deposito="")
            | Q(venda__deposito__isnull=True)
        )
    return qs


def _agg_itens_por_produto(
    desde: datetime, ate: datetime, deposito: str | None = None
) -> list[dict]:
    """
    Soma qtd/valor por produto — mesmo padrão do giro (Sum em colunas reais).
    """
    qs = (
        _qs_itens(desde, ate, deposito=deposito)
        .values("produto_id_externo")
        .annotate(
            qtd=Sum("quantidade"),
            valor=Sum("valor_total"),
        )
        .order_by()
    )
    out: list[dict] = []
    for r in list(qs):
        pid = str(r.get("produto_id_externo") or "").strip()
        if not pid:
            continue
        try:
            qtd = float(r.get("qtd") or 0)
            valor = float(r.get("valor") or 0)
        except (TypeError, ValueError):
            continue
        if qtd <= 0:
            continue
        out.append({"produto_id_externo": pid, "qtd": qtd, "valor": valor})
    return out


def cmv_vendida_de_rows(rows, meta) -> tuple[Decimal, int, int]:
    """Custo cadastro × qtd. Retorna (total, skus_com_custo, skus_sem_custo)."""
    total = Decimal("0")
    skus_ok = 0
    skus_sem = 0
    for r in rows or []:
        pid = str(r.get("produto_id_externo") or "").strip()
        if not pid:
            continue
        try:
            qtd = Decimal(str(r.get("qtd") or 0))
        except Exception:
            continue
        if qtd <= 0:
            continue
        try:
            custo_u = Decimal(str((meta.get(pid) or {}).get("custo") or 0))
        except Exception:
            custo_u = Decimal("0")
        if custo_u <= 0:
            skus_sem += 1
            continue
        total += (custo_u * qtd).quantize(Decimal("0.01"))
        skus_ok += 1
    return total, skus_ok, skus_sem


def custo_mercadoria_vendida(
    data_ini: date,
    data_fim: date,
    *,
    deposito: str | None = None,
) -> dict[str, Any]:
    """Custo cadastro × quantidade vendida no período (todas as SKUs)."""
    por_dia = cmv_vendida_por_dia(data_ini, data_fim, deposito=deposito)
    total = sum(por_dia.values())
    desde = datetime.combine(data_ini, time.min)
    ate = datetime.combine(data_fim, time(23, 59, 59))
    desde, ate = _aware_bounds(desde, ate)
    rows = list(
        _qs_itens(desde, ate, deposito=deposito)
        .values("produto_id_externo")
        .annotate(qtd=Sum("quantidade"))
    )
    pids = [str(r.get("produto_id_externo") or "").strip() for r in rows]
    meta = mapa_produtos_meta(pids)
    _, skus_ok, skus_sem = cmv_vendida_de_rows(rows, meta)
    return {
        "ok": True,
        "total": Decimal(str(round(total, 2))),
        "por_dia": por_dia,
        "skus_com_custo": skus_ok,
        "skus_sem_custo": skus_sem,
        "deposito": deposito or "todas",
    }


def cmv_vendida_por_dia(
    data_ini: date,
    data_fim: date,
    *,
    deposito: str | None = None,
) -> dict[str, float]:
    """CMV vendida (cadastro × qtd) por dia — chave ``YYYY-MM-DD``."""
    from django.db.models.functions import TruncDate

    desde = datetime.combine(data_ini, time.min)
    ate = datetime.combine(data_fim, time(23, 59, 59))
    desde, ate = _aware_bounds(desde, ate)
    rows = list(
        _qs_itens(desde, ate, deposito=deposito)
        .annotate(dia=TruncDate("venda__criado_em"))
        .values("dia", "produto_id_externo")
        .annotate(qtd=Sum("quantidade"))
    )
    by_day: dict[str, list[dict]] = defaultdict(list)
    all_pids: set[str] = set()
    for r in rows:
        dia = r.get("dia")
        if dia is None:
            continue
        pid = str(r.get("produto_id_externo") or "").strip()
        if not pid:
            continue
        k = dia.isoformat() if hasattr(dia, "isoformat") else str(dia)[:10]
        by_day[k].append(r)
        all_pids.add(pid)
    meta = mapa_produtos_meta(list(all_pids))
    out: dict[str, float] = {}
    d = data_ini
    while d <= data_fim:
        k = d.isoformat()
        day_rows = by_day.get(k) or []
        if day_rows:
            total, _, _ = cmv_vendida_de_rows(day_rows, meta)
            out[k] = round(float(total), 2)
        else:
            out[k] = 0.0
        d += timedelta(days=1)
    return out



def _parece_id_mongo_ou_sistema(s: str) -> bool:
    """ObjectId hex (24) ou só dígitos longos do ERP — não é código GM da loja."""
    t = (s or "").strip()
    if not t:
        return False
    if len(t) == 24 and all(c in "0123456789abcdef" for c in t.lower()):
        return True
    if t.isdigit() and len(t) >= 6:
        return True
    return False


def _codigo_gm_preferido(*candidatos: object) -> str:
    """Prioriza código GM (NFe/GM); nunca devolve ObjectId Mongo na coluna Código."""
    humanos: list[str] = []
    for raw in candidatos:
        t = str(raw or "").strip()
        if not t or _parece_id_mongo_ou_sistema(t):
            continue
        if t.upper().startswith("GM"):
            return t
        humanos.append(t)
    return humanos[0] if humanos else ""


def mapa_produtos_meta(pids: list[str]) -> dict[str, dict]:
    """nome, codigo (GM), categoria, marca, custo, comissao_% e comissao_R$."""
    from produtos.catalogo_agro import produto_agro_para_row
    from produtos.models import Produto

    out: dict[str, dict] = {}
    clean = [str(x).strip() for x in pids if str(x).strip()]
    for i in range(0, len(clean), 500):
        slice_ids = clean[i : i + 500]
        for p in Produto.objects.filter(produto_externo_id__in=slice_ids):
            pid = str(p.produto_externo_id or "").strip()
            if not pid:
                continue
            row = produto_agro_para_row(p)
            out[pid] = {
                "nome": (row.get("nome") or p.nome or pid).strip(),
                "codigo": _codigo_gm_preferido(
                    row.get("codigo_nfe"),
                    row.get("codigo_gm"),
                    p.codigo_nfe,
                    row.get("codigo"),
                    p.codigo_interno,
                ),
                "categoria": (row.get("categoria") or p.categoria or "").strip() or "Sem categoria",
                "marca": (row.get("marca") or p.marca or "").strip() or "Sem marca",
                "custo": float(row.get("preco_custo") or p.custo or 0),
                "comissao_pct": None,
                "comissao_rs": None,
            }
    missing = [x for x in clean if x not in out]
    if missing:
        out.update(_mapa_meta_mongo(missing[:800]))
    return out


def _mapa_meta_mongo(pids: list[str]) -> dict[str, dict]:
    try:
        from produtos.views import obter_conexao_mongo
    except Exception:
        return {}
    client, db = obter_conexao_mongo()
    if db is None or client is None:
        return {}
    out: dict[str, dict] = {}
    col = getattr(client, "col_p", "DtoProduto")
    for i in range(0, len(pids), 400):
        slice_ids = pids[i : i + 400]
        try:
            cur = db[col].find(
                {"_id": {"$in": slice_ids}},
                {
                    "Nome": 1,
                    "Codigo": 1,
                    "CodigoInterno": 1,
                    "CodigoNFe": 1,
                    "CodigoNfe": 1,
                    "CodigoGM": 1,
                    "Categoria": 1,
                    "NomeCategoria": 1,
                    "Marca": 1,
                    "NomeMarca": 1,
                    "PrecoCusto": 1,
                    "ValorCusto": 1,
                    "ComissaoVendedor": 1,
                    "ComissaoVendedorPercentual": 1,
                    "PercentualComissao": 1,
                    "_id": 1,
                },
            )
        except Exception as exc:
            logger.warning("relatorios meta mongo: %s", exc)
            continue
        for doc in cur:
            pid = str(doc.get("_id") or "").strip()
            if not pid:
                continue
            try:
                custo = float(doc.get("PrecoCusto") or doc.get("ValorCusto") or 0)
            except (TypeError, ValueError):
                custo = 0.0
            pct = doc.get("ComissaoVendedorPercentual") or doc.get("PercentualComissao")
            rs = doc.get("ComissaoVendedor")
            try:
                pct_f = float(pct) if pct is not None else None
            except (TypeError, ValueError):
                pct_f = None
            try:
                rs_f = float(rs) if rs is not None else None
            except (TypeError, ValueError):
                rs_f = None
            out[pid] = {
                "nome": (doc.get("Nome") or pid).strip(),
                "codigo": _codigo_gm_preferido(
                    doc.get("CodigoNFe"),
                    doc.get("CodigoNfe"),
                    doc.get("CodigoGM"),
                    doc.get("Codigo"),
                    doc.get("CodigoInterno"),
                ),
                "categoria": (
                    doc.get("Categoria") or doc.get("NomeCategoria") or "Sem categoria"
                ).strip()
                or "Sem categoria",
                "marca": (
                    doc.get("Marca") or doc.get("NomeMarca") or "Sem marca"
                ).strip()
                or "Sem marca",
                "custo": custo,
                "comissao_pct": pct_f,
                "comissao_rs": rs_f,
            }
    return out


def ranking_produtos(
    desde: datetime,
    ate: datetime,
    *,
    ordenar: str = "valor",
    sentido: str = "mais",
    limite: int = 100,
) -> list[dict]:
    rows = _agg_itens_por_produto(desde, ate)
    reverse = sentido != "menos"
    key = "qtd" if ordenar == "qtd" else "valor"
    rows.sort(key=lambda x: x[key], reverse=reverse)
    lim = int(limite or 0)
    if lim > 0:
        rows = rows[: max(1, min(50000, lim))]
    pids = [str(r["produto_id_externo"]) for r in rows]
    meta = mapa_produtos_meta(pids)
    out: list[dict] = []
    for i, r in enumerate(rows, start=1):
        pid = str(r["produto_id_externo"])
        m = meta.get(pid) or {}
        qtd = float(r["qtd"] or 0)
        valor = float(r["valor"] or 0)
        out.append(
            {
                "pos": i,
                "produto_id": pid,
                "codigo": m.get("codigo") or "",
                "nome": m.get("nome") or pid,
                "categoria": m.get("categoria") or "Sem categoria",
                "qtd": round(qtd, 3),
                "valor": round(valor, 2),
                "ticket_medio": round(valor / qtd, 2) if qtd else 0.0,
            }
        )
    return out


def vendas_por_grupo(
    desde: datetime, ate: datetime, deposito: str | None = None
) -> list[dict]:
    agg = _agg_itens_por_produto(desde, ate, deposito=deposito)
    pids = [str(r["produto_id_externo"]) for r in agg]
    meta = mapa_produtos_meta(pids)
    buckets: dict[str, dict] = {}
    for r in agg:
        pid = str(r["produto_id_externo"])
        cat = (meta.get(pid) or {}).get("categoria") or "Sem categoria"
        b = buckets.setdefault(cat, {"grupo": cat, "qtd": 0.0, "valor": 0.0, "itens": 0})
        b["qtd"] += float(r["qtd"] or 0)
        b["valor"] += float(r["valor"] or 0)
        b["itens"] += 1
    rows = sorted(buckets.values(), key=lambda x: x["valor"], reverse=True)
    total = sum(x["valor"] for x in rows) or 1.0
    out = []
    for i, r in enumerate(rows, start=1):
        out.append(
            {
                "pos": i,
                "grupo": r["grupo"],
                "qtd": round(r["qtd"], 3),
                "valor": round(r["valor"], 2),
                "skus": r["itens"],
                "pct": round(100.0 * r["valor"] / total, 1),
            }
        )
    return out


def receita_categorias_pdv(
    data_ini: date,
    data_fim: date,
    *,
    deposito: str | None = None,
    top: int = 6,
) -> dict[str, Any]:
    """Faturamento PDV por categoria do cadastro (top + Outros)."""
    desde = datetime.combine(data_ini, time.min)
    ate = datetime.combine(data_fim, time(23, 59, 59))
    desde, ate = _aware_bounds(desde, ate)
    rows = vendas_por_grupo(desde, ate, deposito=deposito)
    total = float(sum(float(r.get("valor") or 0) for r in rows))
    if total <= 0:
        return {"ok": True, "total": 0.0, "fatias": []}
    top_n = max(1, int(top or 6))
    head = rows[:top_n]
    resto = sum(float(r.get("valor") or 0) for r in rows[top_n:])
    fatias: list[dict[str, Any]] = []
    for r in head:
        val = round(float(r.get("valor") or 0), 2)
        if val <= 0:
            continue
        fatias.append(
            {
                "nome": str(r.get("grupo") or "Sem categoria"),
                "valor": val,
                "pct": round(100.0 * val / total, 1),
            }
        )
    if resto > 0.005:
        fatias.append(
            {
                "nome": "Outros",
                "valor": round(resto, 2),
                "pct": round(100.0 * resto / total, 1),
            }
        )
    return {"ok": True, "total": round(total, 2), "fatias": fatias}


def vendas_por_marca(
    desde: datetime,
    ate: datetime,
    *,
    ordenar: str = "valor",
) -> list[dict]:
    """Faturamento e quantidade agrupados pela marca do cadastro (Agro/overlay/Mongo)."""
    agg = _agg_itens_por_produto(desde, ate)
    pids = [str(r["produto_id_externo"]) for r in agg]
    meta = mapa_produtos_meta(pids)
    buckets: dict[str, dict] = {}
    for r in agg:
        pid = str(r["produto_id_externo"])
        marca_raw = (meta.get(pid) or {}).get("marca") or "Sem marca"
        marca = str(marca_raw).strip() or "Sem marca"
        key = marca.casefold()
        b = buckets.setdefault(
            key, {"marca": marca, "qtd": 0.0, "valor": 0.0, "itens": 0}
        )
        b["qtd"] += float(r["qtd"] or 0)
        b["valor"] += float(r["valor"] or 0)
        b["itens"] += 1
    sort_key = "qtd" if ordenar == "qtd" else "valor"
    rows = sorted(buckets.values(), key=lambda x: x[sort_key], reverse=True)
    total_valor = sum(x["valor"] for x in rows) or 1.0
    out = []
    for i, r in enumerate(rows, start=1):
        out.append(
            {
                "pos": i,
                "marca": r["marca"],
                "qtd": round(r["qtd"], 3),
                "valor": round(r["valor"], 2),
                "skus": r["itens"],
                "pct": round(100.0 * r["valor"] / total_valor, 1),
            }
        )
    return out


def curva_abc(
    desde: datetime,
    ate: datetime,
    *,
    todos: bool = False,
    lim_tela: int = 500,
    categoria: str | None = None,
) -> tuple[list[dict], dict]:
    """
    Classifica produtos do período (ou de uma categoria).
    Por padrão mostra só os primeiros ``lim_tela``; com ``todos=True`` lista inteira.
    % e classes usam o faturamento **total** do recorte (período ou categoria).
    """
    rows = ranking_produtos(desde, ate, ordenar="valor", sentido="mais", limite=0)
    categorias = sorted(
        {(r.get("categoria") or "Sem categoria").strip() or "Sem categoria" for r in rows},
        key=lambda x: x.casefold(),
    )
    cat_raw = (categoria or "").strip()
    cat_ativa = ""
    if cat_raw:
        for c in categorias:
            if c.casefold() == cat_raw.casefold():
                cat_ativa = c
                break
        if not cat_ativa:
            cat_ativa = cat_raw
        rows = [
            r
            for r in rows
            if ((r.get("categoria") or "Sem categoria").strip() or "Sem categoria").casefold()
            == cat_ativa.casefold()
        ]
    total_bruto = sum(r["valor"] for r in rows)
    total = total_bruto or 1.0
    acum = 0.0
    out: list[dict] = []
    for i, r in enumerate(rows, start=1):
        acum += r["valor"]
        pct_acum = 100.0 * acum / total
        if pct_acum <= 80.0:
            classe = "A"
        elif pct_acum <= 95.0:
            classe = "B"
        else:
            classe = "C"
        out.append(
            {
                **r,
                "pos": i,
                "pct": round(100.0 * r["valor"] / total, 2),
                "pct_acum": round(pct_acum, 2),
                "classe": classe,
            }
        )
    n_total = len(out)
    lim = max(1, int(lim_tela or 500))
    truncado = (not todos) and n_total > lim
    mostrar = out if todos else out[:lim]
    return mostrar, {
        "total_periodo": round(total_bruto, 2),
        "n_total": n_total,
        "n_tela": len(mostrar),
        "truncado": truncado,
        "todos": bool(todos),
        "categorias": categorias,
        "categoria": cat_ativa,
    }


def margem_produtos(
    desde: datetime,
    ate: datetime,
    *,
    ordenar: str = "margem_rs",
    limite: int = 100,
) -> list[dict]:
    rows = ranking_produtos(desde, ate, ordenar="valor", sentido="mais", limite=limite)
    pids = [r["produto_id"] for r in rows]
    meta = mapa_produtos_meta(pids)
    out = []
    for r in rows:
        m = meta.get(r["produto_id"]) or {}
        custo_u = float(m.get("custo") or 0)
        custo_tot = round(custo_u * r["qtd"], 2)
        margem_rs = round(r["valor"] - custo_tot, 2)
        margem_pct = round(100.0 * margem_rs / r["valor"], 1) if r["valor"] else 0.0
        out.append(
            {
                **r,
                "custo_unit": round(custo_u, 2),
                "custo_total": custo_tot,
                "margem_rs": margem_rs,
                "margem_pct": margem_pct,
            }
        )
    key = "margem_pct" if ordenar == "margem_pct" else "margem_rs"
    out.sort(key=lambda x: x[key], reverse=True)
    for i, r in enumerate(out, start=1):
        r["pos"] = i
    return out


def vendas_por_operador(desde: datetime, ate: datetime) -> list[dict]:
    from django.db.models import Count

    from produtos.models import VendaAgro

    qs = (
        VendaAgro.objects.filter(
            devolvida_em__isnull=True,
            criado_em__gte=desde,
            criado_em__lte=ate,
        )
        .values("usuario_registro")
        .annotate(total=Sum("total"), frete=Sum("frete"), n=Count("id"))
        .order_by("-total")
    )
    out = []
    for i, r in enumerate(qs, start=1):
        op = (r["usuario_registro"] or "").strip() or "(sem operador)"
        total = float(r["total"] or 0)
        frete = float(r["frete"] or 0)
        out.append(
            {
                "pos": i,
                "operador": op,
                "vendas": int(r["n"] or 0),
                "total": round(total, 2),
                "frete": round(frete, 2),
                "ticket": round(total / r["n"], 2) if r["n"] else 0.0,
            }
        )
    return out


def ranking_clientes(
    desde: datetime,
    ate: datetime,
    *,
    ordenar: str = "valor",
    limite: int = 100,
) -> list[dict]:
    from django.db.models import Count
    from produtos.models import VendaAgro

    qs = (
        VendaAgro.objects.filter(
            devolvida_em__isnull=True,
            criado_em__gte=desde,
            criado_em__lte=ate,
        )
        .exclude(Q(cliente_nome="") & Q(cliente_documento="") & Q(cliente_id_erp=""))
        .values("cliente_nome", "cliente_documento", "cliente_id_erp")
        .annotate(total=Sum("total"), n=Count("id"))
    )
    order = "-total" if ordenar != "qtd" else "-n"
    rows = list(qs.order_by(order)[: max(1, min(500, int(limite)))])
    out = []
    for i, r in enumerate(rows, start=1):
        nome = (r["cliente_nome"] or "").strip() or "(sem nome)"
        doc = (r["cliente_documento"] or "").strip()
        total = float(r["total"] or 0)
        n = int(r["n"] or 0)
        out.append(
            {
                "pos": i,
                "cliente": nome,
                "documento": doc,
                "vendas": n,
                "total": round(total, 2),
                "ticket": round(total / n, 2) if n else 0.0,
            }
        )
    return out


def formas_pagamento(desde: datetime, ate: datetime) -> list[dict]:
    from produtos.models import VendaAgro

    buckets: dict[str, float] = {}
    n_por: dict[str, int] = {}
    qs = VendaAgro.objects.filter(
        devolvida_em__isnull=True,
        criado_em__gte=desde,
        criado_em__lte=ate,
    ).only("forma_pagamento", "pagamentos_json", "total")
    for v in qs.iterator(chunk_size=800):
        pagos = v.pagamentos_json
        if isinstance(pagos, list) and pagos:
            for p in pagos:
                if not isinstance(p, dict):
                    continue
                forma = str(p.get("forma") or p.get("Forma") or "Outros").strip() or "Outros"
                try:
                    val = float(p.get("valor") or p.get("Valor") or 0)
                except (TypeError, ValueError):
                    val = 0.0
                buckets[forma] = buckets.get(forma, 0.0) + val
                n_por[forma] = n_por.get(forma, 0) + 1
        else:
            forma = (v.forma_pagamento or "").strip() or "Outros"
            try:
                val = float(v.total or 0)
            except (TypeError, ValueError):
                val = 0.0
            buckets[forma] = buckets.get(forma, 0.0) + val
            n_por[forma] = n_por.get(forma, 0) + 1
    total = sum(buckets.values()) or 1.0
    rows = sorted(buckets.items(), key=lambda x: x[1], reverse=True)
    out = []
    for i, (forma, valor) in enumerate(rows, start=1):
        out.append(
            {
                "pos": i,
                "forma": forma,
                "vendas": n_por.get(forma, 0),
                "total": round(valor, 2),
                "pct": round(100.0 * valor / total, 1),
            }
        )
    return out


def comparativo_periodos(
    desde_a: datetime,
    ate_a: datetime,
    desde_b: datetime,
    ate_b: datetime,
) -> dict[str, Any]:
    from django.db.models import Count
    from produtos.models import VendaAgro

    def _tot(d0, d1):
        agg = VendaAgro.objects.filter(
            devolvida_em__isnull=True,
            criado_em__gte=d0,
            criado_em__lte=d1,
        ).aggregate(total=Sum("total"), n=Count("id"))
        itens = _agg_itens_por_produto(d0, d1)
        qtd = sum(x["qtd"] for x in itens)
        valor = sum(x["valor"] for x in itens)
        return {
            "vendas": int(agg["n"] or 0),
            "faturamento": round(float(agg["total"] or 0), 2),
            "itens": round(float(qtd), 3),
            "itens_rs": round(float(valor), 2),
        }

    a = _tot(desde_a, ate_a)
    b = _tot(desde_b, ate_b)

    def _var(na, nb):
        if not nb:
            return None if not na else 100.0
        return round(100.0 * (na - nb) / nb, 1)

    return {
        "a": a,
        "b": b,
        "var": {
            "vendas": _var(a["vendas"], b["vendas"]),
            "faturamento": _var(a["faturamento"], b["faturamento"]),
            "itens": _var(a["itens"], b["itens"]),
            "itens_rs": _var(a["itens_rs"], b["itens_rs"]),
        },
    }


def comissao_estimada(desde: datetime, ate: datetime, *, limite: int = 200) -> list[dict]:
    rows = ranking_produtos(desde, ate, ordenar="valor", sentido="mais", limite=limite)
    meta = mapa_produtos_meta([r["produto_id"] for r in rows])
    # enriquecer comissão via mongo se faltou
    missing_comissao = [
        r["produto_id"]
        for r in rows
        if (meta.get(r["produto_id"]) or {}).get("comissao_pct") is None
        and (meta.get(r["produto_id"]) or {}).get("comissao_rs") is None
    ]
    if missing_comissao:
        meta.update(_mapa_meta_mongo(missing_comissao[:800]))
    out = []
    for r in rows:
        m = meta.get(r["produto_id"]) or {}
        pct = m.get("comissao_pct")
        rs_u = m.get("comissao_rs")
        com_pct = round(r["valor"] * (float(pct) / 100.0), 2) if pct else 0.0
        com_rs = round(r["qtd"] * float(rs_u), 2) if rs_u else 0.0
        total_com = round(com_pct + com_rs, 2)
        out.append(
            {
                **r,
                "comissao_pct": pct if pct is not None else "",
                "comissao_rs_unit": rs_u if rs_u is not None else "",
                "comissao_estimada": total_com,
            }
        )
    out.sort(key=lambda x: x["comissao_estimada"], reverse=True)
    for i, r in enumerate(out, start=1):
        r["pos"] = i
    return out


def giro_e_parado(*, dias_giro: int = 30, dias_parado: int = 90, limite: int = 150) -> dict:
    from produtos.dashboard_estoque_financeiro_util import (
        obter_estoque_parado_90d,
        obter_top_giro_30d,
    )
    from produtos.models import ItemVendaAgro

    # funções existentes usam 30/90 fixos — reutilizamos e documentamos
    giro = obter_top_giro_30d(limite)
    parado, total_parado = obter_estoque_parado_90d(limite)
    pids = {
        str(x.get("produto_id") or "").strip()
        for x in list(giro) + list(parado)
        if str(x.get("produto_id") or "").strip()
    }
    ultimas_vendas: dict[str, datetime] = {}
    if pids:
        qs_ult = (
            ItemVendaAgro.objects.filter(
                produto_id_externo__in=list(pids),
                venda__devolvida_em__isnull=True,
            )
            .values("produto_id_externo")
            .annotate(ultima_venda=Max("venda__criado_em"))
        )
        ultimas_vendas = {
            str(r["produto_id_externo"] or "").strip(): r["ultima_venda"]
            for r in qs_ult
            if str(r["produto_id_externo"] or "").strip()
        }
    return {
        "giro": [
            {
                "pos": i,
                "produto_id": g["produto_id"],
                "nome": g["nome"],
                "qtd": round(g["total_vendido"], 3),
                "valor": round(g["receita_gerada"], 2),
                "ultima_venda": ultimas_vendas.get(str(g["produto_id"] or "").strip()),
            }
            for i, g in enumerate(giro, start=1)
        ],
        "parado": [
            {
                "pos": i,
                "produto_id": p["produto_id"],
                "nome": p["nome"],
                "estoque": p["estoque_atual"],
                "custo": p["custo"],
                "valor_parado": p["valor_parado"],
                "ultima_venda": ultimas_vendas.get(str(p["produto_id"] or "").strip()),
            }
            for i, p in enumerate(parado, start=1)
        ],
        "total_parado": total_parado,
        "dias_giro": dias_giro,
        "dias_parado": dias_parado,
    }


def ruptura_estoque(*, dias_venda: int = 30, limite: int = 150) -> list[dict]:
    """Produtos com venda recente e saldo C+V zerado."""
    from produtos.dashboard_estoque_financeiro_util import _pids_vendidos_desde
    from produtos.estoque_saldo_agro_util import mapa_saldos_operacionais_agro
    from produtos.views import obter_conexao_mongo

    vendidos = list(_pids_vendidos_desde(dias_venda))
    if not vendidos:
        return []
    client, db = obter_conexao_mongo()
    meta = mapa_produtos_meta(vendidos[:2000])
    out: list[dict] = []
    for i in range(0, len(vendidos), 600):
        slice_ids = vendidos[i : i + 600]
        saldos = mapa_saldos_operacionais_agro(slice_ids, db=db, client=client)
        for pid in slice_ids:
            s = saldos.get(pid) or {}
            estoque = float(s.get("saldo_centro") or 0) + float(s.get("saldo_vila") or 0)
            if estoque > 0.0001:
                continue
            m = meta.get(pid) or {}
            out.append(
                {
                    "produto_id": pid,
                    "codigo": m.get("codigo") or "",
                    "nome": m.get("nome") or pid,
                    "categoria": m.get("categoria") or "Sem categoria",
                    "estoque": 0.0,
                }
            )
            if len(out) >= limite:
                break
        if len(out) >= limite:
            break
    for i, r in enumerate(out, start=1):
        r["pos"] = i
    return out


def montar_xlsx(
    titulo: str,
    headers: list[str],
    rows: list[list[Any]],
    *,
    subtitulo: str = "",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"[:31]
    fill = PatternFill("solid", fgColor="1E293B")
    font_h = Font(bold=True, color="FFFFFF")
    ws.append([titulo])
    ws["A1"].font = Font(bold=True, size=14)
    if subtitulo:
        ws.append([subtitulo])
    ws.append([])
    start = ws.max_row + 1
    ws.append(headers)
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=start, column=col)
        cell.fill = fill
        cell.font = font_h
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(list(row))
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions["A"].width = 8
    if len(headers) > 2:
        ws.column_dimensions["C"].width = 36
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def xlsx_http_response(nome_arquivo: str, content: bytes) -> HttpResponse:
    resp = HttpResponse(
        content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'
    return resp


def fmt_brl(v: float | Decimal | None) -> str:
    try:
        n = float(v or 0)
    except (TypeError, ValueError):
        n = 0.0
    s = f"{n:,.2f}"
    return "R$ " + s.replace(",", "X").replace(".", ",").replace("X", ".")


def fmt_data_curta(v: datetime | date | None) -> str:
    if not v:
        return "-"
    if isinstance(v, datetime):
        try:
            v = timezone.localtime(v)
        except Exception:
            pass
        v = v.date()
    return v.strftime("%d/%m/%Y")

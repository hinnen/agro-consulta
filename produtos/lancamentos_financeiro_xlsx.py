"""
Planilha Excel no layout do relatório financeiro de referência (colunas resumidas + bloco manual).
"""
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

_MESES_ABR = (
    "jan",
    "fev",
    "mar",
    "abr",
    "mai",
    "jun",
    "jul",
    "ago",
    "set",
    "out",
    "nov",
    "dez",
)


def fmt_venc_dd_mmm(data_iso: str | None) -> str:
    if not data_iso:
        return ""
    s = str(data_iso).strip()
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        try:
            y, m, d = int(s[0:4]), int(s[5:7]), int(s[8:10])
            dt = datetime(y, m, d)
            return f"{dt.day:02d}/{_MESES_ABR[dt.month - 1]}"
        except (ValueError, IndexError):
            pass
    return s[:16]


def fmt_brl_pdf(val: float) -> str:
    x = float(val)
    neg = x < 0
    x = abs(x)
    inteiro, frac = f"{x:.2f}".split(".")
    inteiro = f"{int(float(inteiro)):,}".replace(",", ".")
    out = f"R$ {inteiro},{frac}"
    return f"-{out}" if neg else out


def ref_periodo_label(v_de, v_ate) -> str:
    """Texto da célula Período: intervalo quando há duas datas; senão só início, só fim ou hoje."""
    hoje = timezone.localdate()
    if v_de is not None and v_ate is not None:
        return f"{v_de.strftime('%d/%m/%Y')} – {v_ate.strftime('%d/%m/%Y')}"
    if v_de is not None:
        return f"A partir de {v_de.strftime('%d/%m/%Y')}"
    if v_ate is not None:
        return f"Até {v_ate.strftime('%d/%m/%Y')}"
    return hoje.strftime("%d/%m/%Y")


def montar_planilha_financeiro_padrao(
    linhas: list[dict[str, Any]],
    *,
    despesa: bool,
    v_de,
    v_ate,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Financeiro"

    head_font = Font(bold=True)

    ws["A1"] = "Período"
    ws["B1"] = ref_periodo_label(v_de, v_ate)
    ws["A1"].font = head_font

    label_mov = "Pago" if despesa else "Recebido"
    headers = (
        "Vencimento",
        "Cliente / favorecido",
        "Plano conta",
        "Valor bruto",
        label_mov,
        "QUAL CONTA?",
        "Observações",
    )
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = head_font

    row_idx = 3
    for row in linhas:
        vb = float(row.get("valor_bruto") or 0)
        vm = float(row.get("valor_movimentado") or 0)
        pago = bool(row.get("pago"))
        qual_conta = (row.get("banco") or "").strip() or (row.get("forma_pagamento") or "").strip()
        obs_parts = []
        if row.get("descricao"):
            obs_parts.append(str(row.get("descricao") or "").strip())
        if row.get("observacoes"):
            obs_parts.append(str(row.get("observacoes") or "").strip())
        obs = " — ".join([p for p in obs_parts if p])[:500]

        c_plano = (row.get("plano_conta") or "").strip()
        gr = (row.get("grupo") or "").strip()
        if gr:
            c_plano = f"{c_plano} · {gr}".strip(" ·") if c_plano else gr

        if pago and abs(vm) <= 0.005 and abs(vb) > 0.005:
            mov_txt = fmt_brl_pdf(vb)
        elif pago or abs(vm) > 0.005:
            mov_txt = fmt_brl_pdf(vm)
        else:
            mov_txt = ""

        ws.cell(row=row_idx, column=1, value=fmt_venc_dd_mmm(row.get("data_vencimento")))
        ws.cell(row=row_idx, column=2, value=row.get("cliente") or "")
        ws.cell(row=row_idx, column=3, value=c_plano)
        ws.cell(row=row_idx, column=4, value=fmt_brl_pdf(vb))
        ws.cell(row=row_idx, column=5, value=mov_txt)
        ws.cell(row=row_idx, column=6, value=qual_conta)
        ws.cell(row=row_idx, column=7, value=obs)
        for c in range(1, 8):
            ws.cell(row=row_idx, column=c).alignment = Alignment(vertical="top", wrap_text=True)
        row_idx += 1

    row_idx += 2
    for col, title in enumerate(
        ("VALOR ENTRADA", "VALOR DIVIDA", "DATA VENCIMENTO", "BENEFICIARIO"),
        start=1,
    ):
        cell = ws.cell(row=row_idx, column=col, value=title)
        cell.font = head_font
    row_idx += 1
    for _ in range(6):
        row_idx += 1

    row_idx += 1
    ws.cell(row=row_idx, column=1, value="OBSERVAÇÃO").font = head_font
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="ENTROU ALGUM DINHEIRO DE FORA?")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 24
    ws.column_dimensions["G"].width = 40

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()

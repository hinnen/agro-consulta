"""Exportação Excel — histórico de retiradas / saídas do caixa."""
from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

EXPORT_COLS: list[tuple[str, str]] = [
    ("data", "Data"),
    ("hora", "Hora"),
    ("operador_pin", "Operador (PIN)"),
    ("forma", "Forma de pagamento"),
    ("plano", "Plano de contas"),
    ("quem", "Quem levou"),
    ("valor", "Valor (R$)"),
    ("banco", "Conta / banco"),
    ("descricao", "Descrição"),
    ("observacoes", "Observações"),
    ("sessao_id", "Sessão caixa"),
    ("fonte", "Fonte"),
    ("registro_id", "ID registro"),
]

EXPORT_COLS_OBRIGATORIAS = frozenset({"data", "hora", "operador_pin", "forma"})

_COL_MAP = {k: label for k, label in EXPORT_COLS}


def normalizar_colunas_export(raw: str | None) -> list[str]:
    pedidas = [c.strip() for c in str(raw or "").split(",") if c.strip()]
    out: list[str] = []
    vistos: set[str] = set()
    for key in EXPORT_COLS_OBRIGATORIAS:
        if key not in vistos:
            out.append(key)
            vistos.add(key)
    for key in pedidas:
        if key in _COL_MAP and key not in vistos:
            out.append(key)
            vistos.add(key)
    if len(out) == len(EXPORT_COLS_OBRIGATORIAS):
        return [k for k, _ in EXPORT_COLS]
    return out


def headers_export(colunas: list[str] | None) -> list[tuple[str, str]]:
    keys = colunas or [k for k, _ in EXPORT_COLS]
    return [( _COL_MAP[k], k) for k in keys if k in _COL_MAP]


def _fmt_data(d: date | None) -> str:
    if not d:
        return ""
    return d.strftime("%d/%m/%Y")


def _fmt_hora(ts) -> str:
    if not ts:
        return ""
    try:
        return timezone.localtime(ts).strftime("%H:%M")
    except Exception:
        return ""


def linha_export_planilha(row: dict[str, Any]) -> dict[str, Any]:
    op = (row.get("operador_pin") or row.get("operador") or "").strip() or "—"
    return {
        "data": _fmt_data(row.get("data")),
        "hora": _fmt_hora(row.get("criado_em")),
        "operador_pin": op,
        "forma": (row.get("forma") or "").strip() or "—",
        "plano": (row.get("plano") or "").strip() or "—",
        "quem": (row.get("quem") or "").strip() or "—",
        "valor": float(row.get("valor") or 0),
        "banco": (row.get("banco") or "").strip() or "—",
        "descricao": (row.get("descricao") or "").strip(),
        "observacoes": (row.get("observacoes") or "").strip(),
        "sessao_id": row.get("sessao_id") or "",
        "fonte": (row.get("fonte") or "").strip(),
        "registro_id": (row.get("id") or "").strip(),
    }


def montar_xlsx_retiradas(
    rows: list[dict[str, Any]],
    *,
    colunas: list[str] | None = None,
    data_de: date | None = None,
    data_ate: date | None = None,
    plano_filtro: str = "",
    quem_filtro: str = "",
    total: Decimal | None = None,
    truncado: bool = False,
) -> bytes:
    hdrs = headers_export(colunas)
    wb = Workbook()
    ws = wb.active
    ws.title = "Retiradas"

    hdr_fill = PatternFill("solid", fgColor="FFEDD5")
    hdr_font = Font(bold=True, color="9A3412")
    meta_font = Font(bold=True)

    periodo = ""
    if data_de and data_ate:
        periodo = f"{data_de.strftime('%d/%m/%Y')} – {data_ate.strftime('%d/%m/%Y')}"
    elif data_de:
        periodo = f"A partir de {data_de.strftime('%d/%m/%Y')}"
    elif data_ate:
        periodo = f"Até {data_ate.strftime('%d/%m/%Y')}"

    ws["A1"] = "Período"
    ws["B1"] = periodo or timezone.localdate().strftime("%d/%m/%Y")
    ws["A1"].font = meta_font

    filtros_txt = []
    if (plano_filtro or "").strip():
        filtros_txt.append(f"Plano: {plano_filtro.strip()}")
    if (quem_filtro or "").strip():
        filtros_txt.append(f"Quem: {quem_filtro.strip()}")
    if not filtros_txt:
        filtros_txt.append("Plano e quem: todos")
    ws["A2"] = "Filtros"
    ws["B2"] = " · ".join(filtros_txt)
    ws["A2"].font = meta_font

    ws["A3"] = "Registros"
    ws["B3"] = len(rows)
    ws["A3"].font = meta_font

    if total is not None:
        ws["A4"] = "Total (R$)"
        ws["B4"] = float(total)
        ws["B4"].number_format = "#,##0.00"
        ws["A4"].font = meta_font

    if truncado:
        row_aviso = 5
        ws[f"A{row_aviso}"] = "Aviso"
        ws[f"B{row_aviso}"] = "Lista truncada — refine o período ou filtros."
        ws[f"A{row_aviso}"].font = Font(bold=True, color="B45309")
        header_row = row_aviso + 1
    else:
        header_row = 6

    for col, (label, _key) in enumerate(hdrs, start=1):
        c = ws.cell(row=header_row, column=col, value=label)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    for ri, src in enumerate(rows, start=header_row + 1):
        line = linha_export_planilha(src)
        for col, (_label, key) in enumerate(hdrs, start=1):
            val = line.get(key)
            cell = ws.cell(row=ri, column=col)
            if key == "valor":
                cell.value = float(val or 0)
                cell.number_format = "#,##0.00"
            elif key in ("data", "hora", "operador_pin", "forma", "registro_id", "sessao_id"):
                cell.value = str(val) if val is not None else ""
                if key in ("data", "hora"):
                    cell.number_format = "@"
            else:
                cell.value = val if val is not None else ""

    widths = {
        "data": 12,
        "hora": 8,
        "operador_pin": 22,
        "forma": 18,
        "plano": 36,
        "quem": 22,
        "valor": 14,
        "banco": 22,
        "descricao": 42,
        "observacoes": 28,
        "sessao_id": 12,
        "fonte": 12,
        "registro_id": 14,
    }
    for col, (_label, key) in enumerate(hdrs, start=1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = widths.get(key, 16)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
